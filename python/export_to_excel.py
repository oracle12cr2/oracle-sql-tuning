#!/usr/bin/env python3
"""
SQL 튜닝 분석 결과를 Excel로 내보내기

사용법:
    python export_to_excel.py --json output/traces/cj6f6qcqaux57_awr_*.json
    python export_to_excel.py --json output/traces/  # 디렉토리 전체
    python export_to_excel.py --detected output/traces/detected_*.json  # Phase1 감지 결과
"""

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, GradientFill,
                               PatternFill, Side)
from openpyxl.utils import get_column_letter

# ============================================
# 스타일 정의
# ============================================
HEADER_FONT      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_FILL      = PatternFill("solid", start_color="1F3864")
SUB_HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
SUB_HEADER_FILL  = PatternFill("solid", start_color="2E75B6")
TITLE_FONT       = Font(name="Arial", bold=True, size=14, color="1F3864")
LABEL_FONT       = Font(name="Arial", bold=True, size=10, color="1F3864")
NORMAL_FONT      = Font(name="Arial", size=10)
MONO_FONT        = Font(name="Consolas", size=9)
WARN_FILL        = PatternFill("solid", start_color="FFF2CC")
DANGER_FILL      = PatternFill("solid", start_color="FFE0E0")
GOOD_FILL        = PatternFill("solid", start_color="E2EFDA")
ALT_FILL         = PatternFill("solid", start_color="EBF3FB")

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
MEDIUM_BORDER = Border(
    left=Side(style="medium", color="2E75B6"),
    right=Side(style="medium", color="2E75B6"),
    top=Side(style="medium", color="2E75B6"),
    bottom=Side(style="medium", color="2E75B6"),
)

def hcell(ws, row, col, value, width=None):
    """헤더 셀"""
    c = ws.cell(row=row, column=col, value=value)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = THIN_BORDER
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def scell(ws, row, col, value, bold=False, mono=False, fill=None, align="left", wrap=False):
    """일반 데이터 셀"""
    c = ws.cell(row=row, column=col, value=value)
    if mono:
        c.font = MONO_FONT
    elif bold:
        c.font = LABEL_FONT
    else:
        c.font = NORMAL_FONT
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = THIN_BORDER
    if fill:
        c.fill = fill
    return c

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

# ============================================
# Sheet 0: 튜닝 보고서 (Cover / Executive Summary)
# ============================================
def write_cover_page(ws, all_data, detected_list, tkprof_data, data_10053, config_data):
    """📄 튜닝 보고서 표지 + 요약"""
    ws.title = "📄 튜닝 보고서"
    ws.sheet_view.showGridLines = False

    # 컬럼 폭 설정
    for ci, w in enumerate([4, 22, 18, 18, 18, 18, 18, 4], 1):
        set_col_width(ws, ci, w)

    row = 1
    # ── 타이틀 영역 ──
    ws.merge_cells("B1:G1")
    ws.row_dimensions[1].height = 12
    row = 2
    ws.merge_cells("B2:G2")
    c = ws.cell(row=2, column=2, value="Oracle SQL Tuning Analysis Report")
    c.font = Font(name="Arial", bold=True, size=22, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 50

    row = 3
    ws.merge_cells("B3:G3")
    c = ws.cell(row=3, column=2, value="SQL 튜닝 자동화 분석 보고서")
    c.font = Font(name="Arial", bold=True, size=14, color="1F3864")
    c.fill = PatternFill("solid", start_color="EBF3FB")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 35

    # ── 보고서 정보 ──
    row = 5
    info_items = [
        ("보고서 생성일", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("분석 대상 DB", config_data.get("database", {}).get("service_name", "") or config_data.get("database", {}).get("sid", "")),
        ("DB 호스트", config_data.get("database", {}).get("host", "")),
        ("분석 도구", "Oracle SQL Tuning Automation Pipeline v2.0"),
    ]
    for label, value in info_items:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        scell(ws, row, 2, label, bold=True)
        scell(ws, row, 4, value, mono=True)
        ws.row_dimensions[row].height = 20
        row += 1

    row += 1

    # ── 분석 결과 요약 ──
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    c = ws.cell(row=row, column=2, value="  분석 결과 요약")
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="2E75B6")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 28
    row += 1

    # 통계 카드
    total_awr = len(all_data)
    total_detected = len(detected_list)
    total_tkprof = len(tkprof_data)
    total_10053 = len(data_10053)
    total_issues_10053 = sum(len(d.get("issues", [])) for d in data_10053)

    # AWR SQL 중 고위험
    high_elapsed = sum(1 for d in all_data
                       for s in (d.get("stats") or [])
                       if float(s.get("avg_elapsed_sec", 0) or 0) > 5)

    stats_items = [
        ("📊 AWR 분석 SQL", f"{total_awr}건", "Phase 2-4 파이프라인 분석 대상"),
        ("🔍 감지된 느린 SQL", f"{total_detected}건", "Phase 1 임계값 초과 SQL"),
        ("📄 tkprof 분석", f"{total_tkprof}건", "10046 트레이스 분석 결과"),
        ("🔬 10053 분석", f"{total_10053}건", "옵티마이저 트레이스 분석"),
        ("🔴 고위험 SQL (>5초)", f"{high_elapsed}건", "평균 경과시간 5초 초과"),
        ("⚠ 10053 이슈", f"{total_issues_10053}건", "옵티마이저 관련 경고/정보"),
    ]

    headers = ["항목", "건수", "설명"]
    for ci, h in enumerate(headers, 2):
        hcell(ws, row, ci, h, [22, 14, 40][ci-2])
    ws.row_dimensions[row].height = 22
    row += 1

    for label, count, desc in stats_items:
        fill = DANGER_FILL if "고위험" in label and high_elapsed > 0 else (WARN_FILL if "이슈" in label and total_issues_10053 > 0 else None)
        scell(ws, row, 2, label, bold=True, fill=fill)
        scell(ws, row, 3, count, align="center", bold=True, fill=fill)
        scell(ws, row, 4, desc, fill=fill)
        ws.row_dimensions[row].height = 20
        row += 1

    row += 1

    # ── 시트 목록 가이드 ──
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    c = ws.cell(row=row, column=2, value="  시트 구성 안내")
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="2E75B6")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 28
    row += 1

    sheet_guide = [
        ("📄 튜닝 보고서", "보고서 표지 및 전체 요약 (현재 시트)"),
        ("🔌 DB 접속 정보", "분석 대상 DB 연결 정보 및 환경"),
        ("📋 요약", "AWR SQL별 핵심 지표 요약"),
        ("📊 실행계획", "SQL별 실행계획 전문"),
        ("📈 AWR 성능통계", "스냅샷별 상세 성능 지표"),
        ("💡 튜닝 가이드", "자동 분석 기반 튜닝 권장사항"),
        ("🔍 감지된 느린 SQL", "Phase 1 임계값 초과 SQL 목록"),
        ("📄 tkprof 원문", "10046 트레이스 tkprof 전문"),
        ("⏱ Parse-Exec-Fetch", "Parse/Execute/Fetch 단계별 통계"),
        ("⏳ 대기 이벤트", "Wait Event 분석"),
        ("🔖 바인드 변수", "사용된 바인드 변수 값"),
        ("🔬 10053 요약", "10053 옵티마이저 트레이스 요약"),
        ("🛤 접근 경로", "테이블별 Access Path 비용 비교"),
        ("📊 10053 통계", "시스템/테이블/인덱스/컬럼 통계"),
        ("⚠ 10053 이슈", "옵티마이저 관련 이슈 및 권장사항"),
        ("⚙ 옵티마이저 파라미터", "Altered/Default 파라미터"),
    ]

    for ci, h in enumerate(["시트", "설명"], 2):
        hcell(ws, row, ci, h, [22, 55][ci-2])
    ws.row_dimensions[row].height = 22
    row += 1

    for sheet_name, desc in sheet_guide:
        fill = ALT_FILL if row % 2 == 0 else None
        scell(ws, row, 2, sheet_name, bold=True, fill=fill)
        scell(ws, row, 3, desc, fill=fill)
        ws.row_dimensions[row].height = 18
        row += 1

    row += 2
    # 푸터
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    c = ws.cell(row=row, column=2, value="Generated by Oracle SQL Tuning Automation Pipeline — github.com/oracle12cr2/oracle-sql-tuning")
    c.font = Font(name="Arial", size=9, italic=True, color="999999")
    c.alignment = Alignment(horizontal="center")


# ============================================
# DB Live Query
# ============================================
def query_db_info(config_data):
    """DB에 접속하여 인스턴스/데이터베이스/SGA/PGA/파라미터 정보 조회"""
    result = {
        'connected': False,
        'error': None,
        'instance': {},
        'database': {},
        'version': [],
        'sga': [],
        'pga': {},
        'os': {},
        'parameters': {},
        'rac_instances': [],
    }

    try:
        sys.path.insert(0, str(Path(__file__).parent))
        from utils import get_oracle_connection, load_config

        config = load_config()
        conn = get_oracle_connection(config)
        cursor = conn.cursor()
        result['connected'] = True

        # V$INSTANCE
        cursor.execute("""
            SELECT INSTANCE_NUMBER, INSTANCE_NAME, HOST_NAME, VERSION_FULL,
                   STARTUP_TIME, STATUS, DATABASE_STATUS, INSTANCE_ROLE,
                   ACTIVE_STATE, LOGINS, PARALLEL
            FROM V$INSTANCE
        """)
        cols = [d[0] for d in cursor.description]
        row = cursor.fetchone()
        if row:
            result['instance'] = dict(zip(cols, row))

        # V$DATABASE
        cursor.execute("""
            SELECT DBID, NAME, DB_UNIQUE_NAME, CREATED, LOG_MODE,
                   OPEN_MODE, PROTECTION_MODE, DATABASE_ROLE,
                   PLATFORM_NAME, FLASHBACK_ON, FORCE_LOGGING
            FROM V$DATABASE
        """)
        cols = [d[0] for d in cursor.description]
        row = cursor.fetchone()
        if row:
            result['database'] = dict(zip(cols, row))

        # V$VERSION
        cursor.execute("SELECT BANNER_FULL FROM V$VERSION WHERE ROWNUM <= 3")
        result['version'] = [r[0] for r in cursor.fetchall()]

        # V$SGA
        cursor.execute("SELECT NAME, VALUE FROM V$SGA")
        result['sga'] = [(r[0], r[1]) for r in cursor.fetchall()]

        # PGA
        cursor.execute("""
            SELECT NAME, VALUE FROM V$PGASTAT
            WHERE NAME IN ('aggregate PGA target parameter',
                           'aggregate PGA auto target',
                           'total PGA allocated',
                           'total PGA inuse',
                           'maximum PGA allocated')
        """)
        for name, val in cursor.fetchall():
            result['pga'][name] = val

        # OS 정보
        cursor.execute("""
            SELECT STAT_NAME, VALUE FROM V$OSSTAT
            WHERE STAT_NAME IN ('NUM_CPUS', 'NUM_CPU_CORES', 'NUM_CPU_SOCKETS',
                                'PHYSICAL_MEMORY_BYTES', 'IDLE_TIME', 'BUSY_TIME')
        """)
        for name, val in cursor.fetchall():
            result['os'][name] = val

        # 주요 파라미터
        cursor.execute("""
            SELECT NAME, VALUE, ISDEFAULT, DESCRIPTION FROM V$PARAMETER
            WHERE NAME IN (
                'optimizer_mode', 'optimizer_features_enable',
                'db_block_size', 'db_file_multiblock_read_count',
                'pga_aggregate_target', 'sga_target', 'sga_max_size',
                'memory_target', 'memory_max_target',
                'cursor_sharing', 'statistics_level',
                'optimizer_adaptive_plans', 'optimizer_adaptive_statistics',
                'parallel_max_servers', 'parallel_threads_per_cpu',
                'result_cache_mode', 'inmemory_size',
                'undo_tablespace', 'undo_retention',
                'open_cursors', 'session_cached_cursors',
                'db_keep_cache_size', 'db_recycle_cache_size',
                'log_buffer', 'processes', 'sessions'
            )
            ORDER BY NAME
        """)
        for name, val, isdef, desc in cursor.fetchall():
            result['parameters'][name] = {
                'value': val,
                'is_default': isdef == 'TRUE',
                'description': desc
            }

        # RAC 인스턴스 (GV$INSTANCE)
        try:
            cursor.execute("""
                SELECT INST_ID, INSTANCE_NAME, HOST_NAME, STATUS, STARTUP_TIME
                FROM GV$INSTANCE ORDER BY INST_ID
            """)
            cols = [d[0] for d in cursor.description]
            result['rac_instances'] = [dict(zip(cols, r)) for r in cursor.fetchall()]
        except Exception:
            pass

        cursor.close()
        conn.close()

    except Exception as e:
        result['error'] = str(e)
        print(f"  DB query warning: {e}")

    return result




def query_display_cursor(config_data, sql_ids=None):
    """DBMS_XPLAN.DISPLAY_CURSOR로 실행계획 조회
    - SYS, SYSTEM 제외
    - LOCKED 계정 제외
    - 일반 OPEN 계정의 SQL만 대상
    """
    results = {}
    try:
        import sys
        from pathlib import Path
        sys.path.insert(0, str(Path(__file__).parent))
        from utils import get_oracle_connection, load_config
        config = load_config()
        conn = get_oracle_connection(config)
        cursor = conn.cursor()

        # LOCKED 계정 + SYS/SYSTEM 목록 조회
        cursor.execute("""
            SELECT USERNAME FROM DBA_USERS
            WHERE ACCOUNT_STATUS LIKE '%LOCKED%'
               OR USERNAME IN ('SYS', 'SYSTEM', 'DBSNMP', 'OUTLN', 'AUDSYS')
        """)
        exclude_users = set(r[0] for r in cursor)
        print(f"  XPLAN exclude users: {len(exclude_users)} accounts (SYS/SYSTEM/LOCKED)")

        # sql_ids가 없으면 V$SQL에서 일반 계정 Top SQL 조회
        if not sql_ids:
            cursor.execute("""
                SELECT DISTINCT SQL_ID
                FROM V$SQL
                WHERE PARSING_SCHEMA_NAME NOT IN (
                    SELECT USERNAME FROM DBA_USERS
                    WHERE ACCOUNT_STATUS LIKE '%LOCKED%'
                       OR USERNAME IN ('SYS', 'SYSTEM', 'DBSNMP', 'OUTLN', 'AUDSYS')
                )
                AND ELAPSED_TIME > 0
                ORDER BY ELAPSED_TIME DESC
                FETCH FIRST 20 ROWS ONLY
            """)
            sql_ids = [r[0] for r in cursor]
            print(f"  XPLAN target: {len(sql_ids)} SQL IDs from V$SQL (normal users)")
        else:
            # 전달된 sql_ids 중에서도 일반 계정 것만 필터
            placeholders = ','.join([f':s{i}' for i in range(len(sql_ids))])
            bind = {f's{i}': sid for i, sid in enumerate(sql_ids)}
            cursor.execute(f"""
                SELECT DISTINCT SQL_ID FROM V$SQL
                WHERE SQL_ID IN ({placeholders})
                AND PARSING_SCHEMA_NAME NOT IN (
                    SELECT USERNAME FROM DBA_USERS
                    WHERE ACCOUNT_STATUS LIKE '%LOCKED%'
                       OR USERNAME IN ('SYS', 'SYSTEM', 'DBSNMP', 'OUTLN', 'AUDSYS')
                )
            """, bind)
            sql_ids = [r[0] for r in cursor]
            print(f"  XPLAN filtered: {len(sql_ids)} SQL IDs (excluded SYS/SYSTEM/LOCKED)")

        for sql_id in sql_ids:
            try:
                # ALLSTATS LAST: 실제 실행 통계 포함
                cursor.execute(
                    "SELECT * FROM TABLE(DBMS_XPLAN.DISPLAY_CURSOR(:1, NULL, :2))",
                    [sql_id, 'ALLSTATS LAST']
                )
                lines = [row[0] for row in cursor]
                if lines and not any('cannot fetch plan' in l for l in lines[:10]):
                    results[sql_id] = lines
                else:
                    # ALLSTATS 없으면 TYPICAL로 재시도
                    cursor.execute(
                        "SELECT * FROM TABLE(DBMS_XPLAN.DISPLAY_CURSOR(:1, NULL, :2))",
                        [sql_id, 'TYPICAL']
                    )
                    lines = [row[0] for row in cursor]
                    if lines and not any('cannot fetch plan' in l for l in lines[:10]):
                        results[sql_id] = lines
            except Exception as e:
                print(f"  XPLAN {sql_id}: {e}")

        cursor.close()
        conn.close()
    except Exception as e:
        print(f"  XPLAN query error: {e}")
    return results


# ============================================
# Sheet: DB 접속 정보
# ============================================
def write_db_info(ws, config_data, data_10053, all_data, db_live=None):
    """🔌 DB 접속 정보 시트"""
    ws.title = "🔌 DB 접속 정보"
    ws.sheet_view.showGridLines = False

    for ci, w in enumerate([4, 28, 35, 25, 4], 1):
        set_col_width(ws, ci, w)

    row = 1
    # 타이틀
    ws.merge_cells("B1:D1")
    c = ws.cell(row=1, column=2, value="DB 접속 정보 및 환경")
    c.font = Font(name="Arial", bold=True, size=14, color="1F3864")
    c.fill = PatternFill("solid", start_color="EBF3FB")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    row = 3

    # ── 접속 정보 (settings.yaml) ──
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    c = ws.cell(row=row, column=2, value="  Oracle 접속 정보 (settings.yaml)")
    c.font = SUB_HEADER_FONT
    c.fill = SUB_HEADER_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 24
    row += 1

    db_cfg = config_data.get("database", {})
    conn_items = [
        ("호스트 (Host)", db_cfg.get("host", "")),
        ("포트 (Port)", str(db_cfg.get("port", 1521))),
        ("Service Name", db_cfg.get("service_name", "")),
        ("SID", db_cfg.get("sid", "")),
        ("접속 사용자", db_cfg.get("user", "")),
        ("SYSDBA 모드", "Yes" if db_cfg.get("as_sysdba") else "No"),
        ("RAC 인스턴스", ", ".join(db_cfg.get("rac_instances", [])) or "N/A"),
        ("비밀번호 환경변수", db_cfg.get("password_env", "")),
    ]

    hcell(ws, row, 2, "항목", 28)
    hcell(ws, row, 3, "값", 35)
    ws.row_dimensions[row].height = 22
    row += 1

    for label, value in conn_items:
        fill = ALT_FILL if row % 2 == 0 else None
        scell(ws, row, 2, label, bold=True, fill=fill)
        scell(ws, row, 3, value, mono=True, fill=fill)
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1

    # ── SSH 접속 정보 ──
    ssh_cfg = config_data.get("trace", {}).get("ssh", {})
    if ssh_cfg.get("enabled"):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        c = ws.cell(row=row, column=2, value="  SSH 접속 정보 (트레이스 수집용)")
        c.font = SUB_HEADER_FONT
        c.fill = SUB_HEADER_FILL
        ws.row_dimensions[row].height = 24
        row += 1

        ssh_items = [
            ("SSH 호스트", ssh_cfg.get("host", "") or db_cfg.get("host", "") + " (DB호스트 동일)"),
            ("SSH 포트", str(ssh_cfg.get("port", 22))),
            ("SSH 사용자", ssh_cfg.get("user", "")),
            ("인증 방식", ssh_cfg.get("auth_method", "")),
            ("키 경로", ssh_cfg.get("key_path", "")),
        ]

        hcell(ws, row, 2, "항목", 28)
        hcell(ws, row, 3, "값", 35)
        row += 1
        for label, value in ssh_items:
            fill = ALT_FILL if row % 2 == 0 else None
            scell(ws, row, 2, label, bold=True, fill=fill)
            scell(ws, row, 3, value, mono=True, fill=fill)
            ws.row_dimensions[row].height = 18
            row += 1
        row += 1

    # ── 인스턴스 정보 (10053 트레이스에서 추출) ──
    db_info_from_trace = {}
    if data_10053:
        for d in data_10053:
            info = d.get("db_info", {})
            if info:
                db_info_from_trace = info
                break

    if db_info_from_trace:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        c = ws.cell(row=row, column=2, value="  인스턴스 정보 (트레이스에서 추출)")
        c.font = SUB_HEADER_FONT
        c.fill = SUB_HEADER_FILL
        ws.row_dimensions[row].height = 24
        row += 1

        inst_items = [
            ("Oracle 버전", db_info_from_trace.get("oracle_version", "")),
            ("인스턴스 이름", db_info_from_trace.get("instance_name", "")),
            ("데이터베이스 이름", db_info_from_trace.get("database_name", "")),
            ("데이터베이스 역할", db_info_from_trace.get("database_role", "")),
            ("노드 이름", db_info_from_trace.get("node_name", "")),
        ]

        hcell(ws, row, 2, "항목", 28)
        hcell(ws, row, 3, "값", 35)
        row += 1
        for label, value in inst_items:
            fill = ALT_FILL if row % 2 == 0 else None
            scell(ws, row, 2, label, bold=True, fill=fill)
            scell(ws, row, 3, value, mono=True, fill=fill)
            ws.row_dimensions[row].height = 18
            row += 1
        row += 1

    # ── 감지 임계값 ──
    det_cfg = config_data.get("detection", {})
    if det_cfg:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        c = ws.cell(row=row, column=2, value="  감지 임계값 (Phase 1)")
        c.font = SUB_HEADER_FONT
        c.fill = SUB_HEADER_FILL
        ws.row_dimensions[row].height = 24
        row += 1

        det_items = [
            ("경과시간 임계값", f"{det_cfg.get('elapsed_threshold_sec', '')}초"),
            ("Buffer Gets 임계값", f"{det_cfg.get('buffer_gets_threshold', ''):,}"),
            ("Disk Reads 임계값", f"{det_cfg.get('disk_reads_threshold', ''):,}"),
            ("최근 대상 범위", f"{det_cfg.get('recent_minutes', '')}분"),
            ("중복 방지 보관", f"{det_cfg.get('dedup_retention_hours', '')}시간"),
            ("제외 사용자", ", ".join(det_cfg.get("exclude_users", []))),
        ]

        hcell(ws, row, 2, "항목", 28)
        hcell(ws, row, 3, "설정값", 35)
        row += 1
        for label, value in det_items:
            fill = ALT_FILL if row % 2 == 0 else None
            scell(ws, row, 2, label, bold=True, fill=fill)
            scell(ws, row, 3, str(value), mono=True, fill=fill)
            ws.row_dimensions[row].height = 18
            row += 1

    # ── 시스템 통계 (10053에서) ──
    if data_10053:
        sys_stats = data_10053[0].get("system_statistics", {})
        if sys_stats:
            row += 1
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            c = ws.cell(row=row, column=2, value="  시스템 통계 (System Statistics)")
            c.font = SUB_HEADER_FONT
            c.fill = SUB_HEADER_FILL
            ws.row_dimensions[row].height = 24
            row += 1

            hcell(ws, row, 2, "항목", 28)
            hcell(ws, row, 3, "값", 20)
            hcell(ws, row, 4, "기본값", 15)
            row += 1
            stats_type = sys_stats.get("stats_type", "")
            scell(ws, row, 2, "통계 유형", bold=True)
            scell(ws, row, 3, stats_type, mono=True, fill=WARN_FILL if stats_type == "NOWORKLOAD" else GOOD_FILL)
            row += 1
            for name, info in sys_stats.items():
                if name == "stats_type" or not isinstance(info, dict):
                    continue
                fill = WARN_FILL if info.get("is_default") else None
                scell(ws, row, 2, name, bold=True, fill=fill)
                scell(ws, row, 3, info.get("value", ""), mono=True, align="right", fill=fill)
                scell(ws, row, 4, info.get("default", ""), mono=True, align="right", fill=fill)
                ws.row_dimensions[row].height = 18
                row += 1

    # ══════════════════════════════════════════
    # DB Live Query 결과 (V$INSTANCE, V$DATABASE 등)
    # ══════════════════════════════════════════
    if db_live and db_live.get("connected"):
        row += 1

        def _section_header(title):
            nonlocal row
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            c = ws.cell(row=row, column=2, value=f"  {title}")
            c.font = SUB_HEADER_FONT
            c.fill = PatternFill("solid", start_color="1F3864")
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 24
            row += 1

        def _kv_table(items, col_headers=("항목", "값")):
            nonlocal row
            for ci, h in enumerate(col_headers, 2):
                hcell(ws, row, ci, h, [28, 40][ci-2] if len(col_headers) == 2 else [28, 30, 20][ci-2])
            row += 1
            for label, value in items:
                fill = ALT_FILL if row % 2 == 0 else None
                scell(ws, row, 2, label, bold=True, fill=fill)
                val_str = str(value) if value is not None else ""
                scell(ws, row, 3, val_str, mono=True, fill=fill)
                ws.row_dimensions[row].height = 18
                row += 1
            row += 1

        # ── V$INSTANCE ──
        inst = db_live.get("instance", {})
        if inst:
            _section_header("V$INSTANCE (실시간 인스턴스 정보)")
            inst_items = [
                ("Instance Number", inst.get("INSTANCE_NUMBER")),
                ("Instance Name", inst.get("INSTANCE_NAME")),
                ("Host Name", inst.get("HOST_NAME")),
                ("Version", inst.get("VERSION_FULL")),
                ("Startup Time", str(inst.get("STARTUP_TIME", ""))),
                ("Status", inst.get("STATUS")),
                ("Database Status", inst.get("DATABASE_STATUS")),
                ("Instance Role", inst.get("INSTANCE_ROLE")),
                ("Active State", inst.get("ACTIVE_STATE")),
                ("Logins", inst.get("LOGINS")),
                ("Parallel", inst.get("PARALLEL")),
            ]
            _kv_table(inst_items)

        # ── V$DATABASE ──
        db = db_live.get("database", {})
        if db:
            _section_header("V$DATABASE (데이터베이스 정보)")
            db_items = [
                ("DBID", db.get("DBID")),
                ("Database Name", db.get("NAME")),
                ("DB Unique Name", db.get("DB_UNIQUE_NAME")),
                ("Created", str(db.get("CREATED", ""))),
                ("Log Mode", db.get("LOG_MODE")),
                ("Open Mode", db.get("OPEN_MODE")),
                ("Protection Mode", db.get("PROTECTION_MODE")),
                ("Database Role", db.get("DATABASE_ROLE")),
                ("Platform", db.get("PLATFORM_NAME")),
                ("Flashback", db.get("FLASHBACK_ON")),
                ("Force Logging", db.get("FORCE_LOGGING")),
            ]
            _kv_table(db_items)

        # ── V$VERSION ──
        versions = db_live.get("version", [])
        if versions:
            _section_header("V$VERSION (Oracle 버전)")
            for v in versions:
                scell(ws, row, 2, v, mono=True)
                ws.row_dimensions[row].height = 18
                row += 1
            row += 1

        # ── RAC 인스턴스 (GV$INSTANCE) ──
        rac = db_live.get("rac_instances", [])
        if len(rac) > 1:
            _section_header("GV$INSTANCE (RAC 인스턴스 목록)")
            rac_headers = ["Inst#", "Instance Name", "Host Name", "Status", "Startup Time"]
            for ci, h in enumerate(rac_headers, 2):
                hcell(ws, row, ci, h, [8, 16, 20, 10, 20][ci-2])
            row += 1
            for r_inst in rac:
                fill = GOOD_FILL if r_inst.get("STATUS") == "OPEN" else WARN_FILL
                scell(ws, row, 2, r_inst.get("INST_ID"), align="center", fill=fill)
                scell(ws, row, 3, r_inst.get("INSTANCE_NAME"), mono=True, fill=fill)
                scell(ws, row, 4, r_inst.get("HOST_NAME"), mono=True, fill=fill)
                scell(ws, row, 5, r_inst.get("STATUS"), align="center", fill=fill)
                scell(ws, row, 6, str(r_inst.get("STARTUP_TIME", "")), fill=fill)
                ws.row_dimensions[row].height = 18
                row += 1
            row += 1

        # ── V$SGA ──
        sga = db_live.get("sga", [])
        if sga:
            _section_header("V$SGA (메모리 구성)")
            for ci, h in enumerate(("구성 요소", "크기"), 2):
                hcell(ws, row, ci, h, [28, 20][ci-2])
            row += 1
            for name, val in sga:
                fill = ALT_FILL if row % 2 == 0 else None
                size_mb = float(val) / (1024 * 1024) if val else 0
                scell(ws, row, 2, name, bold=True, fill=fill)
                scell(ws, row, 3, f"{size_mb:,.1f} MB", mono=True, align="right", fill=fill)
                ws.row_dimensions[row].height = 18
                row += 1
            row += 1

        # ── PGA ──
        pga = db_live.get("pga", {})
        if pga:
            _section_header("PGA 메모리")
            for ci, h in enumerate(("항목", "크기"), 2):
                hcell(ws, row, ci, h, [35, 20][ci-2])
            row += 1
            for name, val in pga.items():
                fill = ALT_FILL if row % 2 == 0 else None
                size_mb = float(val) / (1024 * 1024) if val else 0
                scell(ws, row, 2, name, bold=True, fill=fill)
                scell(ws, row, 3, f"{size_mb:,.1f} MB", mono=True, align="right", fill=fill)
                ws.row_dimensions[row].height = 18
                row += 1
            row += 1

        # ── OS 정보 ──
        os_info = db_live.get("os", {})
        if os_info:
            _section_header("OS 정보 (V$OSSTAT)")
            os_items = []
            for name, val in os_info.items():
                if "MEMORY" in name:
                    os_items.append((name, f"{float(val) / (1024**3):,.1f} GB"))
                elif "CPU" in name:
                    os_items.append((name, str(int(val))))
                else:
                    os_items.append((name, str(val)))
            _kv_table(os_items)

        # ── 주요 파라미터 (V$PARAMETER) ──
        params = db_live.get("parameters", {})
        if params:
            _section_header("주요 초기화 파라미터 (V$PARAMETER)")
            param_headers = ["파라미터", "값", "기본값?"]
            for ci, h in enumerate(param_headers, 2):
                hcell(ws, row, ci, h, [35, 25, 10][ci-2])
            row += 1
            for name, info in sorted(params.items()):
                is_def = info.get("is_default", True)
                fill = None if is_def else WARN_FILL
                scell(ws, row, 2, name, mono=True, fill=fill)
                # 큰 숫자는 MB로 변환 표시
                val = info.get("value", "")
                if val and val.isdigit() and int(val) > 1048576:
                    val_display = f"{val} ({int(val)/(1024*1024):,.0f} MB)"
                else:
                    val_display = val
                scell(ws, row, 3, val_display, mono=True, fill=fill)
                scell(ws, row, 4, "Y" if is_def else "N (변경됨)", align="center",
                      fill=fill, bold=not is_def)
                ws.row_dimensions[row].height = 18
                row += 1

    elif db_live and db_live.get("error"):
        row += 1
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        c = ws.cell(row=row, column=2, value=f"  DB 접속 실패: {db_live['error']}")
        c.font = Font(name="Arial", bold=True, color="FF0000", size=10)
        c.fill = DANGER_FILL
        ws.row_dimensions[row].height = 24


# ============================================
# Sheet 1: 요약 (Summary)
# ============================================
def write_summary(ws, all_data):
    ws.title = "📋 요약"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "SQL 튜닝 분석 리포트"
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", start_color="EBF3FB")

    ws.merge_cells("A2:H2")
    ws["A2"].value = f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  분석 SQL 수: {len(all_data)}건"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="595959")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["SQL_ID", "SQL 텍스트 (요약)", "Plan Hash Value", "실행계획 줄수",
               "AWR 스냅샷 수", "평균 경과시간(s)", "평균 Buffer Gets", "소스"]
    widths  = [20, 50, 18, 14, 14, 18, 20, 20]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hcell(ws, 3, i, h, w)

    ws.row_dimensions[3].height = 22
    freeze(ws, "A4")

    for r, d in enumerate(all_data, 4):
        fill = ALT_FILL if r % 2 == 0 else None
        sql_preview = (d.get("sql_text") or "")[:80].replace("\n", " ")
        stats = d.get("stats") or []
        avg_elapsed = stats[0].get("avg_elapsed_sec", "") if stats else ""
        avg_buf     = stats[0].get("avg_buffer_gets", "") if stats else ""

        scell(ws, r, 1, d.get("sql_id", ""), mono=True, fill=fill)
        scell(ws, r, 2, sql_preview, fill=fill, wrap=True)
        scell(ws, r, 3, d.get("plan_hash_value", ""), align="center", fill=fill)
        scell(ws, r, 4, len(d.get("plan", [])), align="center", fill=fill)
        scell(ws, r, 5, len(stats), align="center", fill=fill)
        scell(ws, r, 6, avg_elapsed, align="right", fill=fill)
        scell(ws, r, 7, avg_buf, align="right", fill=fill)
        scell(ws, r, 8, d.get("source", "awr"), align="center", fill=fill)
        ws.row_dimensions[r].height = 18

# ============================================
# Sheet 2: 실행계획
# ============================================
def write_plans(ws, all_data):
    ws.title = "📊 실행계획"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 130

    row = 1
    for d in all_data:
        sql_id = d.get("sql_id", "")
        plan   = d.get("plan", [])
        if not plan:
            continue

        # SQL ID 헤더
        c = ws.cell(row=row, column=1, value=f"  SQL_ID: {sql_id}")
        c.font = SUB_HEADER_FONT
        c.fill = SUB_HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1

        # SQL 텍스트
        sql_text = (d.get("sql_text") or "").strip()
        c = ws.cell(row=row, column=1, value=sql_text)
        c.font = Font(name="Consolas", size=10, color="1F3864")
        c.fill = PatternFill("solid", start_color="EBF3FB")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        ws.row_dimensions[row].height = max(32, sql_text.count("\n") * 15 + 15)
        row += 1

        # 실행계획 전체를 하나의 셀에 텍스트 그대로
        # 구분선 길이를 가장 긴 줄에 맞춰 통일
        max_len = max((len(line) for line in plan if line.strip()), default=80)
        normalized = []
        for line in plan:
            # 구분선(-로만 구성)은 max_len만큼 늘려줌
            if line.strip() and re.match(r"^[-]+$", line.strip()):
                normalized.append("-" * max_len)
            else:
                normalized.append(line)
        plan_text = "\n".join(normalized)
        line_count = len(plan)

        c = ws.cell(row=row, column=1, value=plan_text)
        c.font = Font(name="Consolas", size=10, color="000000")
        c.fill = PatternFill("solid", start_color="FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = Border(
            left=Side(style="dashed", color="00AA00"),
            right=Side(style="dashed", color="00AA00"),
            top=Side(style="dashed", color="00AA00"),
            bottom=Side(style="dashed", color="00AA00"),
        )
        ws.row_dimensions[row].height = line_count * 15
        row += 3  # SQL 간 여백

# ============================================
# Sheet 3: AWR 성능 통계
# ============================================
def write_awr_stats(ws, all_data):
    ws.title = "📈 AWR 성능통계"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 22

    headers = ["SQL_ID", "스냅샷 시간", "Plan Hash Value", "실행횟수",
               "평균 경과시간(s)", "평균 CPU(s)", "평균 Buffer Gets",
               "평균 Disk Reads", "평균 Rows"]
    widths  = [20, 18, 18, 12, 18, 14, 20, 18, 14]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hcell(ws, 1, i, h, w)

    freeze(ws, "A2")
    row = 2
    for d in all_data:
        sql_id = d.get("sql_id", "")
        for s in (d.get("stats") or []):
            fill = ALT_FILL if row % 2 == 0 else None
            elapsed = s.get("avg_elapsed_sec") or 0
            buf     = s.get("avg_buffer_gets") or 0
            row_fill = DANGER_FILL if float(elapsed) > 5 else (WARN_FILL if float(elapsed) > 1 else fill)

            scell(ws, row, 1, sql_id, mono=True, fill=row_fill)
            scell(ws, row, 2, s.get("snap_time", ""), fill=row_fill, align="center")
            scell(ws, row, 3, s.get("plan_hash_value", ""), fill=row_fill, align="center")
            scell(ws, row, 4, s.get("executions_delta", ""), fill=row_fill, align="right")
            scell(ws, row, 5, elapsed, fill=row_fill, align="right")
            scell(ws, row, 6, s.get("avg_cpu_sec", ""), fill=row_fill, align="right")
            scell(ws, row, 7, buf, fill=row_fill, align="right")
            scell(ws, row, 8, s.get("avg_disk_reads", ""), fill=row_fill, align="right")
            scell(ws, row, 9, s.get("avg_rows", ""), fill=row_fill, align="right")
            ws.row_dimensions[row].height = 16
            row += 1

    # 색상 범례
    row += 1
    ws.cell(row=row, column=1, value="* 색상 기준").font = LABEL_FONT
    row += 1
    for color, label in [(DANGER_FILL, "평균 경과시간 > 5초"), (WARN_FILL, "평균 경과시간 > 1초")]:
        c = ws.cell(row=row, column=1, value=f"  {label}")
        c.font = NORMAL_FONT
        c.fill = color
        row += 1

# ============================================
# Sheet 4: Phase1 감지 SQL
# ============================================
def write_detected(ws, detected_list):
    ws.title = "🔍 감지된 느린 SQL"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 22

    headers = ["SQL_ID", "Plan Hash", "사용자", "모듈",
               "평균 경과시간(s)", "평균 Buffer Gets", "평균 Disk Reads",
               "실행횟수", "감지시간"]
    widths  = [20, 15, 15, 25, 18, 20, 18, 12, 20]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hcell(ws, 1, i, h, w)

    freeze(ws, "A2")
    for r, s in enumerate(detected_list, 2):
        fill = ALT_FILL if r % 2 == 0 else None
        elapsed = float(s.get("elapsed_time_per_exec", 0) or 0)
        row_fill = DANGER_FILL if elapsed > 10 else (WARN_FILL if elapsed > 3 else fill)

        scell(ws, r, 1, s.get("sql_id", ""), mono=True, fill=row_fill)
        scell(ws, r, 2, s.get("plan_hash_value", ""), fill=row_fill, align="center")
        scell(ws, r, 3, s.get("parsing_user", ""), fill=row_fill)
        scell(ws, r, 4, s.get("module", ""), fill=row_fill)
        scell(ws, r, 5, round(elapsed, 3), fill=row_fill, align="right")
        scell(ws, r, 6, s.get("buffer_gets_per_exec", ""), fill=row_fill, align="right")
        scell(ws, r, 7, s.get("disk_reads_per_exec", ""), fill=row_fill, align="right")
        scell(ws, r, 8, s.get("executions", ""), fill=row_fill, align="right")
        scell(ws, r, 9, s.get("detected_at", ""), fill=row_fill, align="center")
        ws.row_dimensions[r].height = 16


# ============================================
# 튜닝 포인트 분석
# ============================================
def analyze_tuning_points(plan_lines, stats):
    """실행계획과 통계를 분석하여 튜닝 포인트 목록 반환"""
    issues = []
    plan_text = "\n".join(plan_lines)

    def parse_num(s):
        if not s: return 0
        s = str(s).strip().replace(",", "")
        try:
            if s.endswith("K"): return float(s[:-1]) * 1_000
            if s.endswith("M"): return float(s[:-1]) * 1_000_000
            if s.endswith("G"): return float(s[:-1]) * 1_000_000_000
            return float(s)
        except: return 0

    if stats:
        s = stats[0]
        elapsed = float(s.get("avg_elapsed_sec") or 0)
        buf     = parse_num(s.get("avg_buffer_gets") or 0)
        reads   = parse_num(s.get("avg_disk_reads") or 0)
        rows    = parse_num(s.get("avg_rows") or 0)

        if elapsed > 10:
            issues.append(("HIGH", "과도한 실행시간",
                f"평균 경과시간 {elapsed:.2f}초. 즉각적인 튜닝 필요.",
                "실행계획 전체 검토 및 인덱스/조인 방식 재설계 권장"))
        elif elapsed > 3:
            issues.append(("MEDIUM", "느린 실행시간",
                f"평균 경과시간 {elapsed:.2f}초. 튜닝 권장.",
                "인덱스 효율 및 실행계획 검토"))

        if buf > 1_000_000:
            issues.append(("HIGH", "과도한 Buffer Gets",
                f"평균 {buf:,.0f} Buffer Gets. 불필요한 블록 읽기 과다.",
                "인덱스 클러스터링 팩터 확인, 적합한 인덱스 생성 또는 파티셔닝 검토"))
        elif buf > 100_000:
            issues.append(("MEDIUM", "높은 Buffer Gets",
                f"평균 {buf:,.0f} Buffer Gets.",
                "인덱스 선택도(Selectivity) 확인 및 복합 인덱스 검토"))

        if reads > 10_000:
            issues.append(("HIGH", "과도한 Physical Reads",
                f"평균 {reads:,.0f} Disk Reads. 버퍼 캐시 적중률 저하.",
                "자주 사용하는 테이블/인덱스 KEEP 버퍼 풀 적용 검토, SGA 크기 확인"))

        if rows > 0 and buf > 0 and (buf / max(rows, 1)) > 100:
            issues.append(("MEDIUM", "행당 Buffer Gets 과다",
                f"행당 평균 {buf/max(rows,1):.1f} Buffer Gets.",
                "인덱스 Range Scan 비효율 또는 클러스터링 팩터 불량 의심\n인덱스 재생성 또는 복합 인덱스 재설계 검토"))

    if "INDEX FAST FULL SCAN" in plan_text:
        issues.append(("MEDIUM", "INDEX FAST FULL SCAN 감지",
            "인덱스 전체를 스캔하고 있음. 대용량일 경우 비효율.",
            "① WHERE 조건에 적합한 복합 인덱스 생성으로 INDEX RANGE SCAN 유도\n"
            "② SELECT 컬럼이 인덱스에 모두 포함되면 Covering Index 활용\n"
            "③ DISTINCT/ORDER BY 제거 가능 여부 검토"))

    if "TABLE ACCESS FULL" in plan_text or "FULL TABLE SCAN" in plan_text:
        issues.append(("HIGH", "Full Table Scan 감지",
            "테이블 전체를 스캔하고 있음.",
            "① WHERE 조건 컬럼에 인덱스 생성\n"
            "② 조건 컬럼의 선택도(Selectivity) 확인\n"
            "③ 파티션 테이블이라면 파티션 프루닝 여부 확인\n"
            "④ 소규모 테이블은 Full Scan이 더 효율적일 수 있음"))

    if "PARTITION LIST ALL" in plan_text or "PARTITION RANGE ALL" in plan_text:
        issues.append(("MEDIUM", "파티션 전체 스캔 (Pruning 미적용)",
            "파티션 키 조건이 없어 모든 파티션을 스캔 중.",
            "① WHERE 절에 파티션 키 컬럼 조건 추가\n"
            "② 파티션 키 설계 재검토 (현재 쿼리 패턴에 맞게 변경)"))

    if "NESTED LOOPS" in plan_text:
        arows = re.findall(r"NESTED LOOPS[^\n]*\n[^|]*\|[^|]*\|[^|]*\|\s*([\d.KMG]+)\s*\|", plan_text)
        for ar in arows:
            if parse_num(ar) > 100_000:
                issues.append(("MEDIUM", "대용량 Nested Loop Join",
                    f"Nested Loop에서 {ar} 행 처리.",
                    "① Hash Join 또는 Sort Merge Join으로 변경 검토\n"
                    "② 드라이빙 테이블 변경 (/*+ LEADING */ 힌트)\n"
                    "③ 내부 루프 테이블의 인덱스 효율 확인"))
                break

    if re.search(r"SORT\s+(UNIQUE|GROUP BY|JOIN|AGGREGATE)", plan_text):
        issues.append(("LOW", "SORT 연산 발생",
            "정렬 연산이 포함되어 있음.",
            "① 정렬 컬럼에 인덱스 생성으로 SORT 연산 제거 가능\n"
            "② PGA 크기 확인 (work_area_size_policy)\n"
            "③ 불필요한 DISTINCT, GROUP BY 제거 검토"))

    if "HASH JOIN" in plan_text:
        issues.append(("LOW", "Hash Join 사용",
            "Hash Join이 사용 중. 대용량 조인에 일반적으로 적합.",
            "① PGA 메모리 부족 시 Disk Spill 발생 → v$sql_workarea 확인\n"
            "② 소용량 테이블 조인이라면 NL Join이 더 효율적일 수 있음"))

    if not issues:
        issues.append(("GOOD", "특이 사항 없음",
            "주요 튜닝 포인트가 발견되지 않았습니다.",
            "현재 실행계획이 적절한 것으로 보입니다."))

    return issues


def write_tuning_guide(ws, all_data):
    """튜닝 가이드 시트 작성"""
    ws.title = "💡 튜닝 가이드"
    ws.sheet_view.showGridLines = False

    col_widths = [20, 14, 30, 45, 65]
    col_headers = ["SQL_ID", "심각도", "유형", "현상", "튜닝 가이드"]
    for ci, (h, w) in enumerate(zip(col_headers, col_widths), 1):
        hcell(ws, 1, ci, h, w)
    ws.row_dimensions[1].height = 22
    freeze(ws, "A2")

    SEV_FILL = {
        "HIGH":   PatternFill("solid", start_color="FFE0E0"),
        "MEDIUM": PatternFill("solid", start_color="FFF2CC"),
        "LOW":    PatternFill("solid", start_color="E2EFDA"),
        "GOOD":   PatternFill("solid", start_color="E2EFDA"),
    }
    SEV_LABEL = {
        "HIGH":   "🔴 HIGH",
        "MEDIUM": "🟡 MEDIUM",
        "LOW":    "🟢 LOW",
        "GOOD":   "🟢 GOOD",
    }

    row = 2
    for d in all_data:
        sql_id = d.get("sql_id", "")
        issues = analyze_tuning_points(d.get("plan", []), d.get("stats", []))

        for severity, issue_type, symptom, guide in issues:
            fill = SEV_FILL.get(severity)
            line_count = guide.count("\n") + 1

            for ci, val in enumerate([sql_id, SEV_LABEL.get(severity, severity), issue_type, symptom, guide], 1):
                c = ws.cell(row=row, column=ci, value=val)
                c.font = MONO_FONT if ci == 1 else (Font(name="Arial", bold=True, size=10) if ci <= 3 else NORMAL_FONT)
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                c.border = THIN_BORDER
                if fill: c.fill = fill

            ws.row_dimensions[row].height = max(30, line_count * 16 + 8)
            row += 1

        row += 1  # SQL 간 여백

# ============================================
# 메인
# ============================================
def load_awr_jsons(path_pattern):
    """AWR JSON 파일 로드"""
    p = Path(path_pattern)
    files = []
    if p.is_dir():
        files = sorted(p.glob("*_awr_*.json"))
    elif p.is_file():
        files = [p]
    else:
        files = sorted(Path(p.parent).glob(p.name))

    data = []
    for f in files:
        try:
            with open(f, encoding="utf-8") as fp:
                data.append(json.load(fp))
            print(f"  로드: {f.name}")
        except Exception as e:
            print(f"  오류: {f.name} - {e}")
    return data


def load_tkprof_jsons(path_pattern):
    """tkprof JSON 파일 로드"""
    p = Path(path_pattern)
    files = []
    if p.is_dir():
        files = sorted(p.glob("*_tkprof_*.json"))
    elif p.is_file():
        files = [p]
    else:
        files = sorted(Path(p.parent).glob(p.name))

    data = []
    for f in files:
        try:
            with open(f, encoding="utf-8") as fp:
                data.append(json.load(fp))
            print(f"  로드(tkprof): {f.name}")
        except Exception as e:
            print(f"  오류: {f.name} - {e}")
    return data


def write_tkprof_full(ws, tkprof_data):
    """📄 tkprof 원문 시트"""
    ws.title = "📄 tkprof 원문"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 130

    row = 1
    for d in tkprof_data:
        sql_id   = d.get("sql_id", "")
        trc_file = d.get("trc_file", "")
        full_text = d.get("tkprof_full_text", "")
        if not full_text:
            continue

        # 헤더
        c = ws.cell(row=row, column=1, value=f"  SQL_ID: {sql_id}  |  TRC: {Path(trc_file).name}")
        c.font = SUB_HEADER_FONT
        c.fill = SUB_HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1

        # tkprof 전체 텍스트
        line_count = full_text.count("\n") + 1
        c = ws.cell(row=row, column=1, value=full_text)
        c.font = Font(name="Consolas", size=9, color="000000")
        c.fill = PatternFill("solid", start_color="FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = Border(
            left=Side(style="dashed", color="00AA00"),
            right=Side(style="dashed", color="00AA00"),
            top=Side(style="dashed", color="00AA00"),
            bottom=Side(style="dashed", color="00AA00"),
        )
        ws.row_dimensions[row].height = line_count * 14
        row += 3


def write_parse_exec_fetch(ws, tkprof_data):
    """⏱ Parse/Execute/Fetch 통계 시트"""
    ws.title = "⏱ Parse-Exec-Fetch"
    ws.sheet_view.showGridLines = False

    headers = ["SQL_ID", "SQL 텍스트(요약)", "Phase",
               "Count", "CPU(s)", "Elapsed(s)", "Disk", "Query", "Current", "Rows"]
    widths  = [20, 45, 10, 8, 10, 10, 10, 10, 10, 10]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        hcell(ws, 1, ci, h, w)
    ws.row_dimensions[1].height = 22
    freeze(ws, "A2")

    row = 2
    PHASE_FILL = {
        "Parse":   PatternFill("solid", start_color="EBF3FB"),
        "Execute": PatternFill("solid", start_color="E2EFDA"),
        "Fetch":   PatternFill("solid", start_color="FFF2CC"),
        "Total":   PatternFill("solid", start_color="F2F2F2"),
    }

    for d in tkprof_data:
        sql_id = d.get("sql_id", "")
        for stmt in d.get("tkprof_statements", []):
            sql_preview = (stmt.get("sql_text") or "")[:80].replace("\n", " ")
            phases = [
                ("Parse",   stmt.get("parse_count"),   stmt.get("parse_cpu"),   stmt.get("parse_elapsed"),
                            stmt.get("parse_disk"),    stmt.get("parse_query"),  stmt.get("parse_current"), stmt.get("parse_rows")),
                ("Execute", stmt.get("execute_count"), stmt.get("execute_cpu"), stmt.get("execute_elapsed"),
                            stmt.get("execute_disk"),  stmt.get("execute_query"),stmt.get("execute_current"),stmt.get("execute_rows")),
                ("Fetch",   stmt.get("fetch_count"),   stmt.get("fetch_cpu"),   stmt.get("fetch_elapsed"),
                            stmt.get("fetch_disk"),    stmt.get("fetch_query"),  stmt.get("fetch_current"), stmt.get("fetch_rows")),
                ("Total",   "",                        stmt.get("total_cpu"),   stmt.get("total_elapsed"),
                            stmt.get("total_disk"),    stmt.get("total_query"),  stmt.get("total_current"), stmt.get("total_rows")),
            ]
            for phase, count, cpu, elapsed, disk, query, current, rows in phases:
                fill = PHASE_FILL.get(phase)
                scell(ws, row, 1, sql_id,       mono=True, fill=fill)
                scell(ws, row, 2, sql_preview,  fill=fill, wrap=True)
                scell(ws, row, 3, phase,        bold=True, fill=fill, align="center")
                scell(ws, row, 4, count,        fill=fill, align="right")
                scell(ws, row, 5, cpu,          fill=fill, align="right")
                scell(ws, row, 6, elapsed,      fill=fill, align="right")
                scell(ws, row, 7, disk,         fill=fill, align="right")
                scell(ws, row, 8, query,        fill=fill, align="right")
                scell(ws, row, 9, current,      fill=fill, align="right")
                scell(ws, row, 10, rows,        fill=fill, align="right")
                ws.row_dimensions[row].height = 16
                row += 1
            row += 1  # stmt 간 여백


def write_wait_events(ws, tkprof_data):
    """⏳ 대기 이벤트 시트"""
    ws.title = "⏳ 대기 이벤트"
    ws.sheet_view.showGridLines = False

    headers = ["SQL_ID", "Wait Event", "대기 횟수", "최대 대기(s)"]
    widths  = [20, 55, 14, 14]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        hcell(ws, 1, ci, h, w)
    ws.row_dimensions[1].height = 22
    freeze(ws, "A2")

    row = 2
    for d in tkprof_data:
        sql_id = d.get("sql_id", "")
        for stmt in d.get("tkprof_statements", []):
            for w_ev in stmt.get("wait_events", []):
                times = w_ev.get("times_waited", 0)
                max_w = float(w_ev.get("max_wait", 0))
                fill = DANGER_FILL if max_w > 1 else (WARN_FILL if max_w > 0.1 else (ALT_FILL if row % 2 == 0 else None))
                scell(ws, row, 1, sql_id,                   mono=True, fill=fill)
                scell(ws, row, 2, w_ev.get("event", ""),    fill=fill)
                scell(ws, row, 3, times,                    fill=fill, align="right")
                scell(ws, row, 4, max_w,                    fill=fill, align="right")
                ws.row_dimensions[row].height = 16
                row += 1


def write_bind_variables(ws, tkprof_data):
    """🔖 바인드 변수 시트"""
    ws.title = "🔖 바인드 변수"
    ws.sheet_view.showGridLines = False

    headers = ["SQL_ID", "SQL 텍스트(요약)", "위치(Position)", "타입", "값"]
    widths  = [20, 45, 14, 14, 40]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        hcell(ws, 1, ci, h, w)
    ws.row_dimensions[1].height = 22
    freeze(ws, "A2")

    row = 2
    for d in tkprof_data:
        sql_id = d.get("sql_id", "")
        for stmt in d.get("tkprof_statements", []):
            sql_preview = (stmt.get("sql_text") or "")[:80].replace("\n", " ")
            binds = stmt.get("bind_variables", [])
            if not binds:
                continue
            for b in binds:
                fill = ALT_FILL if row % 2 == 0 else None
                scell(ws, row, 1, sql_id,                    mono=True, fill=fill)
                scell(ws, row, 2, sql_preview,               fill=fill, wrap=True)
                scell(ws, row, 3, b.get("position", ""),     fill=fill, align="center")
                scell(ws, row, 4, b.get("type", ""),         fill=fill, align="center")
                scell(ws, row, 5, b.get("value", ""),        mono=True, fill=fill)
                ws.row_dimensions[row].height = 16
                row += 1


def load_detected_jsons(path_pattern):
    """Phase1 감지 결과 JSON 로드"""
    p = Path(path_pattern)
    files = sorted(p.parent.glob(p.name)) if not p.is_dir() else sorted(p.glob("detected_*.json"))
    detected = []
    for f in files:
        try:
            with open(f, encoding="utf-8") as fp:
                d = json.load(fp)
                sql_list = d.get("sql_list", [])
                detected_at = d.get("detected_at", "")
                for s in sql_list:
                    s["detected_at"] = detected_at
                detected.extend(sql_list)
        except Exception as e:
            print(f"  오류: {f.name} - {e}")
    return detected


# ============================================
# Sheet: 10053 Optimizer Trace
# ============================================
def load_10053_data(trace_dir):
    """10053 트레이스 파일을 파싱하여 데이터 로드"""
    import types
    trace_path = Path(trace_dir)
    trc_files = sorted(trace_path.glob("10053_*.trc"))
    if not trc_files:
        return []

    # optimizer_trace 모듈 로드 (DB 의존성 없이)
    ot_path = Path(__file__).parent / "optimizer_trace.py"
    if not ot_path.exists():
        print(f"  optimizer_trace.py 없음: {ot_path}")
        return []

    code = ot_path.read_text(encoding="utf-8")
    code = code.replace("import oracledb", "# import oracledb")
    code = code.replace("from trace_collector import SSHClient", "# skip")
    code = code.replace("from utils import get_oracle_connection, load_config", "def load_config(): return {}")
    mod = types.ModuleType("optimizer_trace")
    exec(compile(code, str(ot_path), "exec"), mod.__dict__)

    results = []
    for f in trc_files:
        try:
            analyzer = mod.OptimizerTraceAnalyzer()
            parsed = analyzer.parse_10053(str(f))
            results.append(parsed)
            print(f"  10053 로드: {f.name} (SQL_ID: {parsed.get('sql_id', '?')})")
        except Exception as e:
            print(f"  10053 오류: {f.name} - {e}")
    return results



def write_xplan_sheet(ws, xplan_data):
    """실행계획 (DBMS_XPLAN) 시트"""
    ws.title = "📋 XPLAN 실행계획"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 140

    row = 1
    for sql_id, lines in xplan_data.items():
        # SQL_ID 헤더
        c = ws.cell(row=row, column=1, value=f"  SQL_ID: {sql_id}  (DBMS_XPLAN.DISPLAY_CURSOR)")
        c.font = SUB_HEADER_FONT
        c.fill = SUB_HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1

        # 실행계획 텍스트
        plan_text = "\n".join(lines)
        line_count = len(lines)
        c = ws.cell(row=row, column=1, value=plan_text)
        c.font = Font(name="Consolas", size=9, color="000000")
        c.fill = PatternFill("solid", start_color="FFFFFF")
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        c.border = Border(
            left=Side(style="dashed", color="2E75B6"),
            right=Side(style="dashed", color="2E75B6"),
            top=Side(style="dashed", color="2E75B6"),
            bottom=Side(style="dashed", color="2E75B6"),
        )
        ws.row_dimensions[row].height = max(line_count * 14, 100)
        row += 3


def write_10053_summary(ws, data_list):
    """🔬 10053 요약 시트"""
    ws.title = "🔬 10053 요약"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "10053 Optimizer Trace Analysis"
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", start_color="EBF3FB")

    ws.merge_cells("A2:J2")
    ws["A2"].value = f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  분석 SQL 수: {len(data_list)}건"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="595959")
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = ["SQL_ID", "SQL 텍스트 (요약)", "테이블", "행 수", "최적 경로",
               "최적 인덱스", "최적 비용", "총 비용", "이슈 수", "인스턴스"]
    widths  = [18, 50, 22, 16, 18, 22, 16, 16, 10, 14]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hcell(ws, 3, i, h, w)
    ws.row_dimensions[3].height = 22
    freeze(ws, "A4")

    for r, d in enumerate(data_list, 4):
        fill = ALT_FILL if r % 2 == 0 else None
        sql_preview = (d.get("sql_text") or "")[:80].replace("\n", " ")

        # 테이블 정보
        tables = list(d.get("base_statistics", {}).keys())
        table_name = tables[0] if tables else ""
        table_rows = d.get("base_statistics", {}).get(table_name, {}).get("rows", 0)

        # 최적 경로
        best_method = ""
        best_index = ""
        best_cost = 0
        if d.get("table_access_paths"):
            best = d["table_access_paths"][0].get("best_access", {})
            best_method = best.get("method", "")
            best_index = best.get("index", "") or ""
            best_cost = best.get("cost", 0)

        total_cost = d.get("best_join_order", {}).get("cost", 0)
        issue_count = len(d.get("issues", []))

        scell(ws, r, 1, d.get("sql_id", ""), mono=True, fill=fill)
        scell(ws, r, 2, sql_preview, fill=fill, wrap=True)
        scell(ws, r, 3, table_name, fill=fill)
        scell(ws, r, 4, f"{table_rows:,}" if table_rows else "", align="right", fill=fill)
        scell(ws, r, 5, best_method, fill=fill, align="center")
        scell(ws, r, 6, best_index, fill=fill)
        scell(ws, r, 7, f"{best_cost:,.2f}" if best_cost else "", align="right", fill=fill)
        scell(ws, r, 8, f"{total_cost:,.2f}" if total_cost else "", align="right", fill=fill)
        issue_fill = DANGER_FILL if issue_count > 2 else (WARN_FILL if issue_count > 0 else fill)
        scell(ws, r, 9, issue_count, align="center", fill=issue_fill)
        scell(ws, r, 10, d.get("db_info", {}).get("instance_name", ""), fill=fill, align="center")
        ws.row_dimensions[r].height = 18


def write_10053_access_paths(ws, data_list):
    """🛤 접근 경로 시트"""
    ws.title = "🛤 접근 경로"
    ws.sheet_view.showGridLines = False

    headers = ["SQL_ID", "테이블", "접근 방법", "인덱스", "비용(Cost)",
               "응답시간(Resp)", "병렬도", "Cost_io", "Cost_cpu", "ix_sel",
               "최적 여부", "비용 비율"]
    widths  = [18, 22, 25, 22, 16, 16, 10, 14, 14, 12, 12, 12]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hcell(ws, 1, i, h, w)
    ws.row_dimensions[1].height = 22
    freeze(ws, "A2")

    row = 2
    for d in data_list:
        sql_id = d.get("sql_id", "")
        for path in d.get("table_access_paths", []):
            table_name = path.get("table_name", "")
            best = path.get("best_access", {})
            best_cost = best.get("cost", 0) if best else 0
            methods = sorted(path.get("access_methods", []), key=lambda x: x.get("cost", float("inf")))

            for i, m in enumerate(methods):
                is_best = (i == 0)
                cost = m.get("cost", 0)
                ratio = cost / best_cost if best_cost > 0 else 0
                row_fill = GOOD_FILL if is_best else (DANGER_FILL if ratio > 5 else None)

                scell(ws, row, 1, sql_id, mono=True, fill=row_fill)
                scell(ws, row, 2, table_name, fill=row_fill)
                scell(ws, row, 3, m.get("method", ""), fill=row_fill)
                scell(ws, row, 4, m.get("index", "") or "-", fill=row_fill)
                scell(ws, row, 5, f"{cost:,.2f}", align="right", fill=row_fill)
                scell(ws, row, 6, f"{m.get('response_time', 0):,.2f}", align="right", fill=row_fill)
                scell(ws, row, 7, m.get("degree", ""), align="center", fill=row_fill)
                cost_io = m.get("cost_io")
                cost_cpu = m.get("cost_cpu")
                ix_sel = m.get("ix_sel")
                scell(ws, row, 8, f"{cost_io:,.6f}" if cost_io is not None else "-", align="right", fill=row_fill)
                scell(ws, row, 9, f"{cost_cpu:,}" if cost_cpu is not None else "-", align="right", fill=row_fill)
                scell(ws, row, 10, f"{ix_sel:.6f}" if ix_sel is not None else "-", align="right", fill=row_fill)
                scell(ws, row, 11, "✅ 최적" if is_best else "", align="center", fill=row_fill)
                scell(ws, row, 12, f"{ratio:.1f}x" if not is_best else "1.0x", align="center", fill=row_fill)
                ws.row_dimensions[row].height = 16
                row += 1

            row += 1  # 테이블 간 여백


def write_10053_stats(ws, data_list):
    """📊 테이블/인덱스/컬럼 통계 시트"""
    ws.title = "📊 10053 통계"
    ws.sheet_view.showGridLines = False

    row = 1

    for d in data_list:
        sql_id = d.get("sql_id", "")

        # SQL ID 헤더
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        c = ws.cell(row=row, column=1, value=f"  SQL_ID: {sql_id}")
        c.font = SUB_HEADER_FONT
        c.fill = SUB_HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1

        # ── 시스템 통계 ──
        sys_stats = d.get("system_statistics", {})
        if sys_stats:
            c = ws.cell(row=row, column=1, value="⚡ 시스템 통계")
            c.font = LABEL_FONT
            row += 1
            sys_headers = ["항목", "값", "단위", "기본값", "기본값 여부"]
            for ci, h in enumerate(sys_headers, 1):
                hcell(ws, row, ci, h, [14, 14, 30, 14, 14][ci-1])
            row += 1
            for name, info in sys_stats.items():
                if name == "stats_type" or not isinstance(info, dict):
                    continue
                fill = WARN_FILL if info.get("is_default") else None
                scell(ws, row, 1, name, bold=True, fill=fill)
                scell(ws, row, 2, info.get("value", ""), align="right", fill=fill)
                scell(ws, row, 3, info.get("unit", ""), fill=fill)
                scell(ws, row, 4, info.get("default", ""), align="right", fill=fill)
                scell(ws, row, 5, "기본값" if info.get("is_default") else "커스텀", align="center", fill=fill)
                ws.row_dimensions[row].height = 16
                row += 1
            row += 1

        # ── 테이블 통계 ──
        base_stats = d.get("base_statistics", {})
        if base_stats:
            c = ws.cell(row=row, column=1, value="📋 테이블 통계")
            c.font = LABEL_FONT
            row += 1
            tbl_headers = ["테이블", "별칭", "행 수", "블록 수", "평균 행 길이", "예상 크기(MB)"]
            for ci, h in enumerate(tbl_headers, 1):
                hcell(ws, row, ci, h, [22, 22, 16, 14, 14, 14][ci-1])
            row += 1
            for tbl, stats in base_stats.items():
                size_mb = (stats["blocks"] * 8192) / (1024 * 1024)
                scell(ws, row, 1, tbl, bold=True)
                scell(ws, row, 2, stats["alias"])
                scell(ws, row, 3, f"{stats['rows']:,}", align="right")
                scell(ws, row, 4, f"{stats['blocks']:,}", align="right")
                scell(ws, row, 5, stats["avg_row_len"], align="right")
                scell(ws, row, 6, f"{size_mb:,.1f}", align="right")
                ws.row_dimensions[row].height = 16
                row += 1
            row += 1

        # ── 인덱스 통계 ──
        idx_stats = d.get("index_statistics", {})
        if idx_stats:
            c = ws.cell(row=row, column=1, value="📑 인덱스 통계")
            c.font = LABEL_FONT
            row += 1
            idx_headers = ["인덱스", "컬럼#", "Levels", "Leaf Blocks", "Distinct Keys",
                          "LB/K", "DB/K", "Clustering Factor", "CF/Rows"]
            for ci, h in enumerate(idx_headers, 1):
                hcell(ws, row, ci, h, [22, 10, 8, 14, 14, 8, 8, 18, 10][ci-1])
            row += 1
            for idx_name, info in idx_stats.items():
                cf_ratio = info["clustering_factor"] / info["num_rows"] if info["num_rows"] > 0 else 0
                fill = DANGER_FILL if cf_ratio > 0.5 else (WARN_FILL if cf_ratio > 0.3 else GOOD_FILL if cf_ratio < 0.1 else None)
                scell(ws, row, 1, idx_name, bold=True, fill=fill)
                scell(ws, row, 2, info["columns"], fill=fill)
                scell(ws, row, 3, info["levels"], align="center", fill=fill)
                scell(ws, row, 4, f"{info['leaf_blocks']:,}", align="right", fill=fill)
                scell(ws, row, 5, f"{info['distinct_keys']:,}", align="right", fill=fill)
                scell(ws, row, 6, f"{info['lb_per_key']:.2f}", align="right", fill=fill)
                scell(ws, row, 7, f"{info['db_per_key']:.2f}", align="right", fill=fill)
                scell(ws, row, 8, f"{info['clustering_factor']:,.0f}", align="right", fill=fill)
                scell(ws, row, 9, f"{cf_ratio:.1%}", align="center", fill=fill)
                ws.row_dimensions[row].height = 16
                row += 1
            row += 1

        # ── 컬럼 통계 ──
        col_stats = d.get("column_statistics", {})
        if col_stats:
            c = ws.cell(row=row, column=1, value="📉 컬럼 통계")
            c.font = LABEL_FONT
            row += 1
            col_headers = ["컬럼", "타입", "NDV", "Nulls", "Density", "AvgLen", "Histogram", "Buckets"]
            for ci, h in enumerate(col_headers, 1):
                hcell(ws, row, ci, h, [22, 14, 12, 10, 14, 10, 12, 10][ci-1])
            row += 1
            for col_name, info in col_stats.items():
                hist = info.get("histogram", "None")
                hist_fill = GOOD_FILL if hist not in ("None", "N/A (no stats)") else WARN_FILL
                scell(ws, row, 1, col_name, bold=True)
                scell(ws, row, 2, info["data_type"])
                scell(ws, row, 3, f"{info['ndv']:,}", align="right")
                scell(ws, row, 4, f"{info['nulls']:,}", align="right")
                scell(ws, row, 5, f"{info['density']:.6f}", align="right")
                scell(ws, row, 6, info["avg_len"], align="right")
                scell(ws, row, 7, hist, align="center", fill=hist_fill)
                scell(ws, row, 8, f"{info['buckets']:,}", align="right")
                ws.row_dimensions[row].height = 16
                row += 1
            row += 1

        # ── 쿼리 변환 ──
        qt_list = d.get("query_transformations", [])
        if qt_list:
            c = ws.cell(row=row, column=1, value="🔄 쿼리 변환 (Query Transformations)")
            c.font = LABEL_FONT
            row += 1
            qt_headers = ["유형", "변환명", "상태", "상세"]
            for ci, h in enumerate(qt_headers, 1):
                hcell(ws, row, ci, h, [10, 30, 14, 60][ci-1])
            row += 1
            for qt in qt_list:
                status = qt.get("status", "")
                fill = GOOD_FILL if status == "applied" else (WARN_FILL if status == "bypassed" else None)
                scell(ws, row, 1, qt.get("type", ""), bold=True, fill=fill)
                scell(ws, row, 2, qt.get("name", ""), fill=fill)
                scell(ws, row, 3, status, align="center", fill=fill)
                scell(ws, row, 4, qt.get("detail", ""), fill=fill, wrap=True)
                ws.row_dimensions[row].height = 16
                row += 1

        row += 2  # SQL 간 여백


def write_10053_issues(ws, data_list):
    """⚠ 10053 이슈 시트"""
    ws.title = "⚠ 10053 이슈"
    ws.sheet_view.showGridLines = False

    headers = ["SQL_ID", "심각도", "유형", "현상", "권장사항"]
    widths  = [18, 12, 24, 55, 65]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hcell(ws, 1, i, h, w)
    ws.row_dimensions[1].height = 22
    freeze(ws, "A2")

    SEV_FILL = {
        "WARNING": WARN_FILL,
        "INFO": GOOD_FILL,
        "CRITICAL": DANGER_FILL,
    }
    SEV_LABEL = {
        "WARNING": "🟡 WARNING",
        "INFO": "🟢 INFO",
        "CRITICAL": "🔴 CRITICAL",
    }

    row = 2
    for d in data_list:
        sql_id = d.get("sql_id", "")
        for issue in d.get("issues", []):
            fill = SEV_FILL.get(issue["severity"], None)
            scell(ws, row, 1, sql_id, mono=True, fill=fill)
            scell(ws, row, 2, SEV_LABEL.get(issue["severity"], issue["severity"]), bold=True, fill=fill, align="center")
            scell(ws, row, 3, issue["type"], bold=True, fill=fill)
            scell(ws, row, 4, issue["message"], fill=fill, wrap=True)
            scell(ws, row, 5, issue["recommendation"], fill=fill, wrap=True)
            ws.row_dimensions[row].height = max(30, issue["recommendation"].count("\n") * 16 + 16)
            row += 1
        row += 1


def write_10053_params(ws, data_list):
    """⚙ 옵티마이저 파라미터 시트"""
    ws.title = "⚙ 옵티마이저 파라미터"
    ws.sheet_view.showGridLines = False

    row = 1
    for d in data_list:
        sql_id = d.get("sql_id", "")

        # SQL ID 헤더
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        c = ws.cell(row=row, column=1, value=f"  SQL_ID: {sql_id}")
        c.font = SUB_HEADER_FONT
        c.fill = SUB_HEADER_FILL
        ws.row_dimensions[row].height = 22
        row += 1

        # 변경된 파라미터
        altered = d.get("optimizer_parameters_altered", {})
        if altered:
            c = ws.cell(row=row, column=1, value="📌 변경된 파라미터 (Altered)")
            c.font = LABEL_FONT
            row += 1
            hcell(ws, row, 1, "파라미터", 45)
            hcell(ws, row, 2, "값", 30)
            row += 1
            for param, value in altered.items():
                scell(ws, row, 1, param, mono=True, fill=WARN_FILL)
                scell(ws, row, 2, value, mono=True, fill=WARN_FILL)
                ws.row_dimensions[row].height = 16
                row += 1
            row += 1

        # 주요 기본 파라미터
        default = d.get("optimizer_parameters_default", {})
        key_params = {
            "optimizer_mode", "optimizer_features_enable", "cpu_count", "active_instance_count",
            "db_file_multiblock_read_count", "pga_aggregate_target", "hash_area_size",
            "sort_area_size", "cursor_sharing", "star_transformation_enabled",
            "parallel_threads_per_cpu", "_optimizer_cost_model", "_b_tree_bitmap_plans",
        }
        key_defaults = {k: v for k, v in default.items() if k in key_params}
        if key_defaults:
            c = ws.cell(row=row, column=1, value="📋 주요 기본 파라미터")
            c.font = LABEL_FONT
            row += 1
            hcell(ws, row, 1, "파라미터", 45)
            hcell(ws, row, 2, "값", 30)
            row += 1
            for param, value in key_defaults.items():
                fill = ALT_FILL if row % 2 == 0 else None
                scell(ws, row, 1, param, mono=True, fill=fill)
                scell(ws, row, 2, value, mono=True, fill=fill)
                ws.row_dimensions[row].height = 16
                row += 1

        row += 2


def main():
    parser = argparse.ArgumentParser(description="SQL 튜닝 분석 결과 Excel 내보내기")
    parser.add_argument("--json", type=str, default="./output/traces/",
                        help="AWR JSON 파일 또는 디렉토리 경로")
    parser.add_argument("--detected", type=str, default=None,
                        help="Phase1 감지 결과 JSON 파일 또는 디렉토리")
    parser.add_argument("--output", type=str, default=None,
                        help="출력 엑셀 파일 경로")
    parser.add_argument("--10053", dest="trace_10053", type=str, default=None,
                        help="10053 트레이스 파일 디렉토리 (예: output/traces/)")
    parser.add_argument("--db-password", dest="db_password", type=str, default=None,
                        help="DB 비밀번호 (환경변수 대신 직접 전달)")
    parser.add_argument("--config", type=str, default=None)
    args = parser.parse_args()

    # 출력 경로
    if args.output:
        out_path = Path(args.output)
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = Path("./output/reports") / f"sql_tuning_report_{ts}.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"AWR JSON 로드 중: {args.json}")
    json_path    = args.json
    det_path     = args.detected
    all_data     = load_awr_jsons(json_path)
    tkprof_data  = load_tkprof_jsons(json_path if Path(json_path).is_dir() else str(Path(json_path).parent))

    detected_list = []
    if det_path:
        print(f"감지 결과 로드 중: {det_path}")
        detected_list = load_detected_jsons(det_path)

    # 10053 데이터 로드
    data_10053 = []
    trace_10053_path = args.trace_10053
    if not trace_10053_path:
        # 기본: output/traces/ 에서 자동 탐색
        default_trace_dir = Path(json_path) if Path(json_path).is_dir() else Path(json_path).parent
        if list(default_trace_dir.glob("10053_*.trc")):
            trace_10053_path = str(default_trace_dir)
    if trace_10053_path:
        print(f"10053 트레이스 로드 중: {trace_10053_path}")
        data_10053 = load_10053_data(trace_10053_path)

    if not all_data and not detected_list and not tkprof_data and not data_10053:
        print("데이터 없음. 종료.")
        sys.exit(1)

    # settings.yaml 로드
    import yaml
    config_data = {}
    config_path = Path(args.config) if args.config else Path(json_path).parent.parent / "config" / "settings.yaml"
    if not config_path.exists():
        config_path = Path("config/settings.yaml")
    if config_path.exists():
        with open(config_path, encoding="utf-8") as fp:
            config_data = yaml.safe_load(fp) or {}
        print(f"  config: {config_path}")

    print(f"\nExcel generating: {out_path}")
    wb = Workbook()
    wb.remove(wb.active)

    # 맨 앞 시트: 튜닝 보고서 + DB 접속 정보
    ws_cover   = wb.create_sheet("📄 튜닝 보고서")
    ws_dbinfo  = wb.create_sheet("🔌 DB 접속 정보")
    # DB Live Query
    # Windows 환경변수 전달 문제 우회: --db-password 옵션 지원
    if hasattr(args, 'db_password') and args.db_password:
        import os
        pwd_env = config_data.get("database", {}).get("password_env", "ORACLE_TUNING_PWD")
        os.environ[pwd_env] = args.db_password
    print("DB live query...")
    db_live = query_db_info(config_data)
    if db_live.get("connected"):
        print(f"  DB connected: {db_live['instance'].get('INSTANCE_NAME', '?')} @ {db_live['instance'].get('HOST_NAME', '?')}")
    elif db_live.get("error"):
        print(f"  DB connection failed: {db_live['error'][:80]}")

    # XPLAN 조회 (V$SQL에 있는 SQL_ID들)
    xplan_data = {}
    if db_live and db_live.get("connected"):
        # AWR 데이터에서 SQL_ID 추출
        xplan_sql_ids = list(set(
            [d.get("sql_id") for d in all_data if d.get("sql_id")] +
            [d.get("sql_id") for d in data_10053 if d.get("sql_id")]
        ))
        if xplan_sql_ids:
            print(f"XPLAN query: {len(xplan_sql_ids)} SQL IDs...")
            xplan_data = query_display_cursor(config_data, xplan_sql_ids)
            print(f"  XPLAN found: {len(xplan_data)} plans")

    write_cover_page(ws_cover, all_data, detected_list, tkprof_data, data_10053, config_data)
    write_db_info(ws_dbinfo, config_data, data_10053, all_data, db_live=db_live)

    # XPLAN 시트
    if xplan_data:
        ws_xplan = wb.create_sheet("📋 XPLAN 실행계획")
        write_xplan_sheet(ws_xplan, xplan_data)

    # 기존 시트
    ws_summary = wb.create_sheet("📋 요약")
    ws_plan    = wb.create_sheet("📊 실행계획")
    ws_awr     = wb.create_sheet("📈 AWR 성능통계")
    ws_tuning  = wb.create_sheet("💡 튜닝 가이드")
    ws_detect  = wb.create_sheet("🔍 감지된 느린 SQL")
    ws_tkprof  = wb.create_sheet("📄 tkprof 원문")
    ws_pef     = wb.create_sheet("⏱ Parse-Exec-Fetch")
    ws_wait    = wb.create_sheet("⏳ 대기 이벤트")
    ws_bind    = wb.create_sheet("🔖 바인드 변수")

    if all_data:
        write_summary(ws_summary, all_data)
        write_plans(ws_plan, all_data)
        write_awr_stats(ws_awr, all_data)
        write_tuning_guide(ws_tuning, all_data)

    write_detected(ws_detect, detected_list)

    if tkprof_data:
        write_tkprof_full(ws_tkprof, tkprof_data)
        write_parse_exec_fetch(ws_pef, tkprof_data)
        write_wait_events(ws_wait, tkprof_data)
        write_bind_variables(ws_bind, tkprof_data)

    # 10053 시트
    if data_10053:
        ws_10053_sum    = wb.create_sheet("🔬 10053 요약")
        ws_10053_access = wb.create_sheet("🛤 접근 경로")
        ws_10053_stats  = wb.create_sheet("📊 10053 통계")
        ws_10053_issues = wb.create_sheet("⚠ 10053 이슈")
        ws_10053_params = wb.create_sheet("⚙ 옵티마이저 파라미터")
        write_10053_summary(ws_10053_sum, data_10053)
        write_10053_access_paths(ws_10053_access, data_10053)
        write_10053_stats(ws_10053_stats, data_10053)
        write_10053_issues(ws_10053_issues, data_10053)
        write_10053_params(ws_10053_params, data_10053)

    wb.save(str(out_path))
    print(f"\n[OK] Saved: {out_path}")
    return str(out_path)


if __name__ == "__main__":
    main()
