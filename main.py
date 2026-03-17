#!/usr/bin/env python3
"""
Oracle SQL Tuning Automation - 통합 실행기 (main.py)
Windows / Linux 양쪽에서 실행 가능

사용법:
    # Phase 1: 느린 SQL 감지
    python main.py detect
    python main.py detect --dry-run

    # Phase 2: 트레이스 수집 (Phase 1 결과 자동 연계)
    python main.py trace
    python main.py trace --sql-id abc123def456

    # Phase 3: tkprof 분석
    python main.py analyze
    python main.py analyze --file output/traces/abc123.trc

    # Phase 4: 리포트 생성
    python main.py report --daily
    python main.py report --weekly

    # 전체 파이프라인 (Phase 1 → 2 → 3 자동 연계)
    python main.py run

    # Windows 스케줄러 등록
    python main.py install-schedule
"""

import argparse
import json
import os
import platform
import subprocess
import sys
import time
import tempfile
import shutil
from datetime import datetime
from pathlib import Path

# 프로젝트 루트 설정
PROJECT_ROOT = Path(__file__).resolve().parent
PYTHON_DIR = PROJECT_ROOT / "python"
sys.path.insert(0, str(PYTHON_DIR))

from utils import load_config, setup_logger


def get_config_path():
    return str(PROJECT_ROOT / "config" / "settings.yaml")


# ============================================
# Phase 1: 느린 SQL 감지
# ============================================
def cmd_detect(args):
    """Phase 1 실행"""
    from slow_sql_detector import detect_slow_sql, filter_duplicates, \
        print_summary, save_detection_result, DedupStore

    # DedupStore import 수정
    from utils import DedupStore

    config = load_config(get_config_path())
    logger = setup_logger("phase1_detector", config)

    try:
        results = detect_slow_sql(config, logger)

        dedup_store = DedupStore(
            config["paths"]["dedup_db"],
            retention_hours=config["detection"]["dedup_retention_hours"],
        )
        dedup_store.cleanup()

        if not args.dry_run:
            results = filter_duplicates(results, dedup_store, logger)

        print_summary(results, logger)

        if results and not args.dry_run:
            output_file = save_detection_result(
                results, config["paths"]["trace_output"], logger
            )
            for r in results:
                dedup_store.mark_collected(
                    r["sql_id"],
                    str(r.get("plan_hash_value", "0")),
                    elapsed_time_sec=r.get("elapsed_sec_per_exec"),
                )
            logger.info(f"Phase 1 완료: {len(results)}건 감지")
            return str(output_file)
        else:
            logger.info("Phase 1 완료: 신규 감지 없음" if not args.dry_run else "Phase 1 완료 (dry-run)")
            return None

    except Exception as e:
        logger.error(f"Phase 1 오류: {e}", exc_info=True)
        return None


# ============================================
# Phase 2: 트레이스 수집
# ============================================
def cmd_trace(args):
    """Phase 2 실행"""
    from trace_collector import TraceCollector, process_detected_file

    config = load_config(get_config_path())
    logger = setup_logger("phase2_collector", config)

    try:
        if args.sql_id:
            logger.info(f"단일 SQL 트레이스: {args.sql_id}")
            collector = TraceCollector(config, logger)
            collector.connect()
            result = collector.process_sql(args.sql_id, wait_seconds=args.wait)
            collector.close()
            logger.info(f"결과: {result['status']}")
            return [result] if result["status"] == "collected" else []

        elif args.file:
            return process_detected_file(args.file, config, logger)

        else:
            # 최근 감지 파일 자동 탐색
            trace_dir = Path(config["paths"]["trace_output"])
            detected_files = sorted(trace_dir.glob("detected_*.json"), reverse=True)
            if detected_files:
                latest = detected_files[0]
                logger.info(f"최근 감지 파일 사용: {latest.name}")
                return process_detected_file(str(latest), config, logger)
            else:
                logger.info("처리할 감지 파일 없음")
                return []

    except Exception as e:
        logger.error(f"Phase 2 오류: {e}", exc_info=True)
        return []


# ============================================
# Phase 3: tkprof 분석
# ============================================
def cmd_analyze(args):
    """Phase 3 실행"""
    from tkprof_analyzer import process_trace_file

    config = load_config(get_config_path())
    logger = setup_logger("phase3_tkprof", config)

    try:
        if args.file:
            trace_path = Path(args.file)
        else:
            trace_path = Path(config["paths"]["trace_output"])

        if trace_path.is_dir():
            trc_files = sorted(trace_path.glob("*.trc"))
            if not trc_files:
                logger.info("분석할 트레이스 파일 없음")
                return

            # 이미 분석된 파일 제외
            tkprof_dir = Path(config["paths"]["tkprof_output"])
            analyzed = {f.stem for f in tkprof_dir.glob("*.prf")} if tkprof_dir.exists() else set()
            new_files = [f for f in trc_files if f.stem not in analyzed]

            if not new_files:
                logger.info("모든 트레이스 파일이 이미 분석됨")
                return

            logger.info(f"Phase 3: {len(new_files)}개 트레이스 분석")
            for trc in new_files:
                process_trace_file(str(trc), config, logger)

        elif trace_path.is_file():
            process_trace_file(str(trace_path), config, logger)

        else:
            logger.error(f"경로 없음: {trace_path}")

    except Exception as e:
        logger.error(f"Phase 3 오류: {e}", exc_info=True)


# ============================================
# Phase 4: 리포트 생성
# ============================================
def cmd_report(args):
    """Phase 4 실행"""
    from report_generator import main as report_main

    # argparse 인자를 report_generator에 전달
    sys.argv = ["report_generator.py", "--config", get_config_path()]
    if args.weekly:
        sys.argv.append("--weekly")
    else:
        sys.argv.append("--daily")
    if args.date:
        sys.argv.extend(["--date", args.date])

    report_main()


# ============================================
# 전체 파이프라인 실행
# ============================================

def cmd_target(args):
    """특정 SQL_ID 타겟 분석 (10046 + tkprof + 10053 + Excel)"""
    config = load_config(get_config_path())
    logger = setup_logger("pipeline", config)

    sql_id = args.sql_id
    db_password = getattr(args, 'db_password', None)
    skip_10046 = getattr(args, 'skip_10046', False)
    skip_10053 = getattr(args, 'skip_10053', False)
    skip_excel = getattr(args, 'skip_excel', False)

    # DB 비밀번호 설정
    if db_password:
        pwd_env = config.get("database", {}).get("password_env", "ORACLE_TUNING_PWD")
        os.environ[pwd_env] = db_password
        config["database"]["password"] = db_password

    # 출력 폴더: output/reports/YYYY.MM.DD_SQL_ID/
    report_dir = Path(config["paths"]["report_output"]) / f"{datetime.now().strftime('%Y.%m.%d')}_{sql_id}"
    report_dir.mkdir(parents=True, exist_ok=True)

    logger.info("=" * 60)
    logger.info(f"타겟 분석: SQL_ID = {sql_id}")
    logger.info(f"출력 폴더: {report_dir}")
    logger.info("=" * 60)

    # XPLAN 실행계획 저장
    try:
        from utils import get_oracle_connection
        conn = get_oracle_connection(config)
        cursor = conn.cursor()
        cursor.execute(
            "SELECT * FROM TABLE(DBMS_XPLAN.DISPLAY_CURSOR(:1, NULL, :2))",
            [sql_id, 'ALLSTATS LAST']
        )
        xplan_lines = [row[0] for row in cursor]
        if xplan_lines and not any('cannot fetch plan' in l for l in xplan_lines[:10]):
            xplan_file = report_dir / f"xplan_{sql_id}.txt"
            with open(xplan_file, 'w', encoding='utf-8') as f:
                f.write('\n'.join(xplan_lines))
            logger.info(f"XPLAN 저장: {xplan_file}")
        else:
            logger.info("XPLAN: 커서에 플랜 없음 (flush됨)")
        cursor.close()
        conn.close()
    except Exception as e:
        logger.warning(f"XPLAN 조회 실패: {e}")

    # SQL 텍스트 조회
    sql_text = None
    try:
        from utils import get_oracle_connection
        conn = get_oracle_connection(config)
        cursor = conn.cursor()
        cursor.execute("SELECT SQL_FULLTEXT FROM V$SQL WHERE SQL_ID = :1 AND ROWNUM = 1", [sql_id])
        row = cursor.fetchone()
        if row:
            sql_text = str(row[0]) if hasattr(row[0], 'read') else str(row[0])
            logger.info(f"SQL: {sql_text[:100]}...")
        else:
            logger.warning(f"V$SQL에서 SQL_ID {sql_id} 못 찾음 (이미 flush됐을 수 있음)")
        cursor.close()
        conn.close()
    except Exception as e:
        logger.warning(f"SQL 텍스트 조회 실패: {e}")

    # --- 10046 트레이스 수집 ---
    trc_file = None
    if not skip_10046:
        logger.info("\n" + "=" * 40)
        logger.info("Step 1: 10046 트레이스 수집")
        logger.info("=" * 40)
        try:
            args.file = None
            args.wait = None
            trace_results = cmd_trace(args)
            collected = [r for r in trace_results if r.get("status") == "collected"]
            if collected:
                trc_file = collected[0].get("trace_file")
                logger.info(f"  수집 완료: {trc_file}")
                # 트레이스 파일을 분석 폴더에 복사
                if trc_file:
                    import shutil
                    try:
                        shutil.copy2(trc_file, str(report_dir))
                        logger.info(f"  복사: {report_dir}")
                    except Exception:
                        pass
            else:
                logger.warning("  트레이스 수집 실패 (계속 진행)")
        except Exception as e:
            logger.warning(f"  10046 오류 (계속 진행): {e}")

    # --- tkprof 분석 ---
    if trc_file:
        logger.info("\n" + "=" * 40)
        logger.info("Step 2: tkprof 분석")
        logger.info("=" * 40)
        try:
            args.file = trc_file
            cmd_analyze(args)
        except Exception as e:
            logger.warning(f"  tkprof 오류 (계속 진행): {e}")

    # --- 10053 옵티마이저 트레이스 ---
    if not skip_10053:
        logger.info("\n" + "=" * 40)
        logger.info("Step 3: 10053 옵티마이저 트레이스")
        logger.info("=" * 40)
        try:
            from optimizer_trace import OptimizerTraceCollector, OptimizerTraceAnalyzer
            collector = OptimizerTraceCollector(config, logger)

            if sql_text:
                trc_10053 = collector.collect_10053(sql_id, sql_text)
                if trc_10053:
                    logger.info(f"  10053 수집: {trc_10053}")
                    analyzer = OptimizerTraceAnalyzer(config, logger)
                    parsed = analyzer.parse_10053(trc_10053)
                    if parsed:
                        report_file = str(report_dir / f"10053_{sql_id}.html")
                        analyzer.generate_10053_report(parsed, report_file)
                        logger.info(f"  10053 리포트: {report_file}")
            else:
                logger.warning("  SQL 텍스트 없어서 10053 스킵")
        except Exception as e:
            logger.warning(f"  10053 오류 (계속 진행): {e}")

    # --- Excel 리포트 ---
    if not skip_excel:
        logger.info("\n" + "=" * 40)
        logger.info("Step 4: Excel 통합 리포트")
        logger.info("=" * 40)
        try:
            import subprocess
            excel_cmd = [
                sys.executable, 'python/export_to_excel.py',
                '--json', str(Path(config["paths"]["trace_output"])),
                '--10053', str(Path(config["paths"]["trace_output"])),
                '--output', str(report_dir / f"tuning_report_{sql_id}.xlsx"),
            ]
            if db_password:
                excel_cmd.extend(['--db-password', db_password])
            result = subprocess.run(excel_cmd, capture_output=True, text=True, timeout=120)
            if result.returncode == 0:
                for line in result.stdout.split('\n'):
                    if '[OK]' in line:
                        logger.info(f"  {line.strip()}")
                        break
            else:
                logger.warning(f"  Excel 오류: {result.stderr[:200]}")
        except Exception as e:
            logger.warning(f"  Excel 오류 (계속 진행): {e}")

    logger.info("\n" + "=" * 60)
    logger.info(f"타겟 분석 완료: SQL_ID = {sql_id}")
    logger.info(f"리포트: {report_dir}")
    logger.info("=" * 60)


def cmd_run(args):
    """Phase 1 → 2 → 3 → 4 → 10053 → Excel 전체 파이프라인"""
    config = load_config(get_config_path())
    logger = setup_logger("pipeline", config)

    logger.info("=" * 60)
    logger.info("전체 파이프라인 실행 시작")
    logger.info("=" * 60)

    # DB 비밀번호를 환경변수에 설정 (Windows SSH 환경변수 전달 문제 우회)
    db_password = getattr(args, 'db_password', None)
    if db_password:
        pwd_env = config.get("database", {}).get("password_env", "ORACLE_TUNING_PWD")
        os.environ[pwd_env] = db_password
        config["database"]["password"] = db_password

    skip_phase1 = getattr(args, 'skip_detect', False)
    skip_10053 = getattr(args, 'skip_10053', False)
    skip_excel = getattr(args, 'skip_excel', False)
    detected_file = None
    trace_results = []
    collected = []
    detected_sql_ids = []

    # --- Phase 1: 느린 SQL 감지 ---
    if not skip_phase1:
        logger.info("\n" + "=" * 40)
        logger.info("Phase 1: 느린 SQL 감지")
        logger.info("=" * 40)
        args.dry_run = False
        detected_file = cmd_detect(args)

        if not detected_file:
            logger.info("감지된 SQL 없음.")
        else:
            # 감지된 SQL ID 목록 추출
            try:
                import json
                with open(detected_file, encoding='utf-8') as f:
                    det_data = json.load(f)
                detected_sql_ids = [s.get('sql_id') for s in det_data.get('sql_list', []) if s.get('sql_id')]
                logger.info(f"감지된 SQL: {len(detected_sql_ids)}건 - {detected_sql_ids}")
            except Exception as e:
                logger.warning(f"감지 결과 파싱 실패: {e}")

    # --- Phase 2: 트레이스 수집 ---
    if detected_file:
        logger.info("\n" + "=" * 40)
        logger.info("Phase 2: 10046 트레이스 수집")
        logger.info("=" * 40)
        args.sql_id = None
        args.file = detected_file
        args.wait = None
        trace_results = cmd_trace(args)
        collected = [r for r in trace_results if r.get("status") == "collected"]
        logger.info(f"수집 완료: {len(collected)}건")

    # --- Phase 3: tkprof 분석 ---
    if collected:
        logger.info("\n" + "=" * 40)
        logger.info("Phase 3: tkprof 분석")
        logger.info("=" * 40)
        for result in collected:
            if result.get("trace_file"):
                args.file = result["trace_file"]
                cmd_analyze(args)

    # --- Phase 4: HTML 리포트 생성 ---
    logger.info("\n" + "=" * 40)
    logger.info("Phase 4: HTML 리포트 생성")
    logger.info("=" * 40)
    try:
        args.daily = True
        args.weekly = False
        args.date = None
        cmd_report(args)
    except Exception as e:
        logger.warning(f"HTML 리포트 생성 중 오류 (계속 진행): {e}")
    except SystemExit:
        logger.warning("Phase 4 리포트 모듈이 종료됨 (계속 진행)")

    # --- Phase 5: 10053 옵티마이저 트레이스 ---
    if not skip_10053 and detected_sql_ids:
        logger.info("\n" + "=" * 40)
        logger.info("Phase 5: 10053 옵티마이저 트레이스")
        logger.info("=" * 40)
        try:
            from optimizer_trace import OptimizerTraceCollector, OptimizerTraceAnalyzer
            collector = OptimizerTraceCollector(config, logger)

            for sql_id in detected_sql_ids:
                try:
                    logger.info(f"10053 수집: {sql_id}")
                    # SQL 텍스트 조회
                    from utils import get_oracle_connection
                    conn = get_oracle_connection(config)
                    cursor = conn.cursor()
                    cursor.execute("SELECT SQL_FULLTEXT FROM V$SQL WHERE SQL_ID = :1 AND ROWNUM = 1",
                                  [sql_id])
                    row = cursor.fetchone()
                    sql_text = str(row[0]) if row else None
                    cursor.close()
                    conn.close()

                    if not sql_text:
                        logger.warning(f"  SQL_ID {sql_id}: V$SQL에서 SQL 텍스트 못 찾음, 스킵")
                        continue

                    trc_file = collector.collect_10053(sql_id, sql_text)
                    if trc_file:
                        logger.info(f"  10053 수집 완료: {trc_file}")
                        # 분석 + HTML 리포트
                        analyzer = OptimizerTraceAnalyzer(config, logger)
                        parsed = analyzer.parse_10053(trc_file)
                        if parsed:
                            report_dir = os.path.join(config["paths"]["report_output"], "10053")
                            os.makedirs(report_dir, exist_ok=True)
                            report_file = os.path.join(report_dir,
                                f"10053_{sql_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html")
                            analyzer.generate_10053_report(parsed, report_file)
                            logger.info(f"  10053 리포트: {report_file}")
                except Exception as e:
                    logger.warning(f"  10053 {sql_id} 오류 (계속 진행): {e}")
        except Exception as e:
            logger.warning(f"10053 Phase 오류 (계속 진행): {e}")

    # --- Phase 6: Excel 통합 리포트 ---
    if not skip_excel:
        logger.info("\n" + "=" * 40)
        logger.info("Phase 6: Excel 통합 리포트 생성")
        logger.info("=" * 40)
        try:
            excel_args = [
                'python/export_to_excel.py',
                '--json', str(Path(config["paths"]["trace_output"])),
                '--10053', str(Path(config["paths"]["trace_output"])),
                '--output', str(Path(config["paths"]["report_output"]) /
                    f"sql_tuning_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"),
            ]
            if db_password:
                excel_args.extend(['--db-password', db_password])

            # subprocess로 export_to_excel.py 실행
            import subprocess
            result = subprocess.run(
                [sys.executable, 'python/export_to_excel.py',
                 '--json', str(Path(config["paths"]["trace_output"])),
                 '--10053', str(Path(config["paths"]["trace_output"])),
                 '--output', str(Path(config["paths"]["report_output"]) /
                     f"sql_tuning_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"),
                ] + (['--db-password', db_password] if db_password else []),
                capture_output=True, text=True, timeout=120
            )
            if result.returncode == 0:
                # 출력에서 파일 경로 추출
                for line in result.stdout.split('\n'):
                    if '[OK]' in line:
                        logger.info(f"  {line.strip()}")
                        break
                else:
                    logger.info("  Excel 리포트 생성 완료")
            else:
                logger.warning(f"  Excel 생성 오류: {result.stderr[:200]}")
        except Exception as e:
            logger.warning(f"Excel 리포트 오류 (계속 진행): {e}")

    logger.info("\n" + "=" * 60)
    logger.info("전체 파이프라인 완료!")
    logger.info("=" * 60)
    logger.info(f"  감지: {len(detected_sql_ids)}건")
    logger.info(f"  트레이스 수집: {len(collected)}건")
    logger.info(f"  리포트: {config['paths']['report_output']}")
    logger.info("=" * 60)


# ============================================
# 파일 정리
# ============================================
def cmd_cleanup(args):
    """오래된 파일 정리"""
    config = load_config(get_config_path())
    logger = setup_logger("cleanup", config)

    trace_days = config["trace"].get("retention_days", 30)
    report_days = 90

    dirs_and_days = [
        (config["paths"]["trace_output"], "*.trc", trace_days),
        (config["paths"]["trace_output"], "detected_*.json", trace_days),
        (config["paths"]["trace_output"], "trace_results_*.json", trace_days),
        (config["paths"]["tkprof_output"], "*.prf", trace_days),
        (config["paths"]["report_output"], "*.html", report_days),
    ]

    now = datetime.now()
    total_removed = 0

    for dir_path, pattern, max_days in dirs_and_days:
        target_dir = Path(dir_path)
        if not target_dir.exists():
            continue

        for f in target_dir.glob(pattern):
            age_days = (now - datetime.fromtimestamp(f.stat().st_mtime)).days
            if age_days > max_days:
                f.unlink()
                total_removed += 1
                logger.info(f"  삭제: {f.name} ({age_days}일 경과)")

    logger.info(f"정리 완료: {total_removed}개 파일 삭제")


# ============================================
# Windows 작업 스케줄러 등록
# ============================================
def cmd_install_schedule(args):
    """Windows 작업 스케줄러 또는 Linux cron 등록"""
    config = load_config(get_config_path())
    python_exe = sys.executable
    main_py = str(PROJECT_ROOT / "main.py")

    if platform.system() == "Windows":
        print("Windows 작업 스케줄러 등록")
        print("=" * 50)

        # Phase 1: 5분 간격 감지
        task_name_detect = "OracleTuning_Detect"
        cmd_detect_task = (
            f'schtasks /create /tn "{task_name_detect}" '
            f'/tr "\\"{python_exe}\\" \\"{main_py}\\" run" '
            f'/sc minute /mo 5 /f'
        )

        # Phase 4: 일간 리포트 매일 07:00
        task_name_daily = "OracleTuning_DailyReport"
        cmd_daily = (
            f'schtasks /create /tn "{task_name_daily}" '
            f'/tr "\\"{python_exe}\\" \\"{main_py}\\" report --daily" '
            f'/sc daily /st 07:00 /f'
        )

        # Phase 4: 주간 리포트 매주 월요일 08:00
        task_name_weekly = "OracleTuning_WeeklyReport"
        cmd_weekly = (
            f'schtasks /create /tn "{task_name_weekly}" '
            f'/tr "\\"{python_exe}\\" \\"{main_py}\\" report --weekly" '
            f'/sc weekly /d MON /st 08:00 /f'
        )

        # 파일 정리: 매일 자정
        task_name_cleanup = "OracleTuning_Cleanup"
        cmd_cleanup_task = (
            f'schtasks /create /tn "{task_name_cleanup}" '
            f'/tr "\\"{python_exe}\\" \\"{main_py}\\" cleanup" '
            f'/sc daily /st 00:00 /f'
        )

        tasks = [
            (task_name_detect, cmd_detect_task, "5분 간격 SQL 감지 + 트레이스"),
            (task_name_daily, cmd_daily, "일간 리포트 (매일 07:00)"),
            (task_name_weekly, cmd_weekly, "주간 리포트 (매주 월 08:00)"),
            (task_name_cleanup, cmd_cleanup_task, "파일 정리 (매일 00:00)"),
        ]

        if args.remove:
            for name, _, desc in tasks:
                print(f"  제거: {name}")
                os.system(f'schtasks /delete /tn "{name}" /f 2>nul')
            print("\n작업 스케줄러 등록 제거 완료")
            return

        print(f"\nPython: {python_exe}")
        print(f"Script: {main_py}\n")

        for name, cmd, desc in tasks:
            print(f"  등록: {name} ({desc})")
            ret = os.system(cmd)
            if ret != 0:
                print(f"    ⚠ 실패 (관리자 권한이 필요할 수 있습니다)")
            else:
                print(f"    ✓ 성공")

        print(f"\n등록 확인:")
        print(f'  schtasks /query /tn "OracleTuning_Detect"')
        print(f"\n제거:")
        print(f"  python main.py install-schedule --remove")

    else:
        # Linux cron
        print("Linux crontab 등록")
        print("=" * 50)

        cron_lines = [
            f"# Oracle SQL Tuning Automation",
            f"*/5 * * * * {python_exe} {main_py} run >> {config['paths']['log_dir']}/cron.log 2>&1",
            f"0 7 * * * {python_exe} {main_py} report --daily >> {config['paths']['log_dir']}/cron.log 2>&1",
            f"0 8 * * 1 {python_exe} {main_py} report --weekly >> {config['paths']['log_dir']}/cron.log 2>&1",
            f"0 0 * * * {python_exe} {main_py} cleanup >> {config['paths']['log_dir']}/cron.log 2>&1",
        ]

        print("\n아래 내용을 crontab -e로 등록하세요:\n")
        for line in cron_lines:
            print(f"  {line}")
        print()


# ============================================
# 상태 확인
# ============================================
def cmd_status(args):
    """현재 상태 확인"""
    config = load_config(get_config_path())

    print("Oracle SQL Tuning Automation - 상태")
    print("=" * 50)
    print(f"OS: {platform.system()} {platform.release()}")
    print(f"Python: {sys.version.split()[0]}")
    print(f"프로젝트: {PROJECT_ROOT}")

    # Oracle 드라이버
    try:
        import oracledb
        print(f"DB 드라이버: oracledb {oracledb.__version__}")
    except ImportError:
        try:
            import cx_Oracle
            print(f"DB 드라이버: cx_Oracle {cx_Oracle.__version__}")
        except ImportError:
            print("DB 드라이버: ⚠ 미설치 (pip install oracledb)")

    # DB 접속 정보 표시
    db = config["database"]
    service_name = db.get("service_name", "")
    sid = db.get("sid", "")
    print(f"\nDB 설정:")
    print(f"  Host: {db['host']}:{db['port']}")
    if service_name:
        print(f"  Service Name: {service_name}")
    if sid:
        print(f"  SID: {sid}")
    if not service_name and not sid:
        print(f"  ⚠ SID/Service Name 미설정!")
    print(f"  User: {db['user']}")
    print(f"  SYSDBA: {'Yes' if db.get('as_sysdba') else 'No'}")
    pwd_set = "설정됨" if db.get("password") else "미설정"
    print(f"  패스워드: {pwd_set} (환경변수: {db.get('password_env', 'ORACLE_TUNING_PWD')})")

    # 실제 접속 테스트
    print(f"\nDB 접속 테스트:")
    try:
        from utils import get_oracle_connection
        conn = get_oracle_connection(config)
        cursor = conn.cursor()

        cursor.execute("""
            SELECT i.instance_name, i.host_name, i.version,
                   i.status, d.name AS db_name, i.startup_time,
                   d.open_mode, d.log_mode,
                   (SELECT value FROM v$parameter WHERE name = 'db_unique_name') AS db_unique_name,
                   (SELECT value FROM v$parameter WHERE name = 'service_names') AS service_names
            FROM v$instance i, v$database d
        """)
        row = cursor.fetchone()
        if row:
            inst_name, host_name, version, status, db_name, startup, \
                open_mode, log_mode, db_unique_name, service_names = row
            print(f"  ✓ 접속 성공!")
            print(f"  DB Name: {db_name}")
            print(f"  DB Unique Name: {db_unique_name}")
            print(f"  Instance: {inst_name}")
            print(f"  DB Host: {host_name}")
            print(f"  Version: {version}")
            print(f"  Status: {status}")
            print(f"  Open Mode: {open_mode}")
            print(f"  Log Mode: {log_mode}")
            print(f"  Service Names: {service_names}")
            print(f"  Startup: {startup}")

            # 설정값과 실제 인스턴스 일치 확인
            config_service = db.get("service_name", "").upper()
            config_sid = db.get("sid", "").upper()
            actual_instance = inst_name.upper()
            actual_db = db_name.upper()
            actual_services = (service_names or "").upper()

            match_found = False
            if config_service:
                if config_service in actual_services or config_service == actual_db:
                    print(f"\n  ✓ 서비스 확인: 설정({db.get('service_name')}) → DB({db_name}) 일치")
                    match_found = True
                else:
                    print(f"\n  ⚠ 서비스 불일치!")
                    print(f"    설정 Service Name: {db.get('service_name')}")
                    print(f"    실제 Service Names: {service_names}")
                    print(f"    실제 DB Name: {db_name}")
            elif config_sid:
                if config_sid == actual_instance or config_sid == actual_db:
                    print(f"\n  ✓ 인스턴스 확인: 설정({db.get('sid')}) = 실제({inst_name}) 일치")
                    match_found = True
                else:
                    print(f"\n  ⚠ 인스턴스 불일치!")
                    print(f"    설정 SID: {db.get('sid')}")
                    print(f"    실제 Instance: {inst_name}")
                    print(f"    실제 DB Name: {db_name}")
                    print(f"    → settings.yaml의 sid 또는 service_name 값을 확인하세요")

        # RAC 환경 확인
        try:
            cursor.execute("""
                SELECT inst_id, instance_name, host_name, status
                FROM gv$instance ORDER BY inst_id
            """)
            rac_rows = cursor.fetchall()
            if len(rac_rows) > 1:
                print(f"\n  RAC 환경 ({len(rac_rows)}개 인스턴스):")
                for inst_id, iname, hname, istatus in rac_rows:
                    current = " ← 현재 접속" if iname == inst_name else ""
                    print(f"    [{inst_id}] {iname} @ {hname} ({istatus}){current}")
        except Exception:
            pass  # GV$ 권한 없을 수 있음

        cursor.close()
        conn.close()
    except Exception as e:
        print(f"  ✗ 접속 실패: {e}")

    # tkprof
    tkprof_path = config["tkprof"]["binary_path"]
    if platform.system() == "Windows":
        tkprof_path = tkprof_path.replace("/", "\\")
        if not tkprof_path.endswith(".exe"):
            tkprof_path += ".exe"
    tkprof_exists = shutil.which("tkprof") or os.path.exists(tkprof_path)
    print(f"tkprof: {'✓ 사용 가능' if tkprof_exists else '⚠ 미발견 (' + tkprof_path + ')'}")

    # SSH 접속 테스트
    ssh_conf = config["trace"].get("ssh", {})
    if ssh_conf.get("enabled"):
        ssh_host = ssh_conf.get("host") or config["database"]["host"]
        ssh_user = ssh_conf.get("user", "oracle")
        ssh_port = ssh_conf.get("port", 22)
        print(f"\nSSH 설정:")
        print(f"  Host: {ssh_user}@{ssh_host}:{ssh_port}")
        print(f"  인증: {ssh_conf.get('auth_method', 'key')}")

        try:
            from trace_collector import SSHClient
            ssh = SSHClient(config, type("L", (), {"info": lambda s, m: None, "warning": lambda s, m: None})())
            ok, msg = ssh.test_connection()
            if ok:
                print(f"  ✓ SSH 접속 성공: {msg}")
            else:
                print(f"  ✗ SSH 접속 실패: {msg}")
        except Exception as e:
            print(f"  ✗ SSH 테스트 오류: {e}")
    else:
        print(f"\nSSH: 비활성화 (settings.yaml → trace.ssh.enabled: true로 변경)")

    # 디렉토리
    print(f"\n디렉토리:")
    for name, path in [
        ("traces", config["paths"]["trace_output"]),
        ("tkprof", config["paths"]["tkprof_output"]),
        ("reports", config["paths"]["report_output"]),
        ("logs", config["paths"]["log_dir"]),
    ]:
        p = Path(path)
        if p.exists():
            count = len(list(p.iterdir()))
            print(f"  {name}: {p} ({count}개 파일)")
        else:
            print(f"  {name}: {p} (미생성)")

    # 최근 실행 이력
    from utils import DedupStore
    dedup_path = config["paths"]["dedup_db"]
    if Path(dedup_path).exists():
        store = DedupStore(dedup_path)
        recent = store.get_recent(hours=24)
        print(f"\n최근 24시간 감지: {len(recent)}건")
        for sql_id, plan_hash, collected_at, elapsed in recent[:5]:
            print(f"  {sql_id} | elapsed={elapsed}s | {collected_at}")

    # Windows 스케줄러 상태
    if platform.system() == "Windows":
        print(f"\nWindows 작업 스케줄러:")
        for task in ["OracleTuning_Detect", "OracleTuning_DailyReport", "OracleTuning_WeeklyReport"]:
            ret = os.system(f'schtasks /query /tn "{task}" >nul 2>&1')
            status = "✓ 등록됨" if ret == 0 else "✗ 미등록"
            print(f"  {task}: {status}")


# ============================================
# Export: 엑셀 내보내기
# ============================================
def cmd_export(args):
    """AWR JSON → Excel 내보내기"""
    from export_to_excel import load_awr_jsons, load_tkprof_jsons, load_detected_jsons, \
        write_summary, write_plans, write_awr_stats, write_tuning_guide, write_detected, \
        write_tkprof_full, write_parse_exec_fetch, write_wait_events, write_bind_variables
    from openpyxl import Workbook

    config = load_config(get_config_path())
    logger = setup_logger("export", config)

    # 경로 기본값: 프로젝트 루트 기준
    json_path = args.json or str(PROJECT_ROOT / config["paths"]["trace_output"])
    det_path  = args.detected or None
    out_path  = Path(args.output) if args.output else \
                PROJECT_ROOT / config["paths"]["report_output"] / \
                f"sql_tuning_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    out_path.parent.mkdir(parents=True, exist_ok=True)

    logger.info(f"AWR JSON 로드: {json_path}")
    all_data    = load_awr_jsons(json_path)
    tkprof_data = load_tkprof_jsons(json_path if Path(json_path).is_dir() else str(Path(json_path).parent))

    detected_list = []
    if det_path:
        logger.info(f"감지 결과 로드: {det_path}")
        detected_list = load_detected_jsons(det_path)

    if not all_data and not detected_list and not tkprof_data:
        logger.error("내보낼 데이터 없음")
        return

    wb = Workbook()
    wb.remove(wb.active)
    ws_summary = wb.create_sheet("📋 요약")
    ws_plan    = wb.create_sheet("📊 실행계획")
    ws_awr     = wb.create_sheet("📈 AWR 성능통계")
    ws_tuning  = wb.create_sheet("💡 튜닝 가이드")
    ws_detect  = wb.create_sheet("🔍 감지된 느린 SQL")
    ws_tkprof  = wb.create_sheet("📄 tkprof 원문")
    ws_pef     = wb.create_sheet("⏱ Parse-Exec-Fetch")
    ws_wait    = wb.create_sheet("⏳ 대기 이벤트")
    ws_bind    = wb.create_sheet("🔖 바인드 변수")

    if all_data:
        write_summary(ws_summary, all_data)
        write_plans(ws_plan, all_data)
        write_awr_stats(ws_awr, all_data)
        write_tuning_guide(ws_tuning, all_data)
    write_detected(ws_detect, detected_list)
    if tkprof_data:
        write_tkprof_full(ws_tkprof, tkprof_data)
        write_parse_exec_fetch(ws_pef, tkprof_data)
        write_wait_events(ws_wait, tkprof_data)
        write_bind_variables(ws_bind, tkprof_data)

    wb.save(str(out_path))
    logger.info(f"Excel 저장 완료: {out_path}")
    print(f"\n✅ Excel 저장: {out_path}")


# ============================================
# 메인 파서
# ============================================

# ============================================
# 10053 옵티마이저 트레이스
# ============================================
def cmd_optimizer_trace(args):
    """10053 트레이스 수집"""
    from optimizer_trace import OptimizerTraceCollector
    config = load_config(get_config_path())
    logger = setup_logger("optimizer_trace", config)

    if not args.sql_id:
        logger.error("--sql-id 필수")
        return

    try:
        collector = OptimizerTraceCollector(config, logger)

        # SQL 텍스트 조회
        sql_text = None
        if not args.sql_text:
            from utils import get_oracle_connection
            conn = get_oracle_connection(config)
            cursor = conn.cursor()
            cursor.execute("SELECT SQL_FULLTEXT FROM V$SQL WHERE SQL_ID = :1 AND ROWNUM = 1",
                          [args.sql_id])
            row = cursor.fetchone()
            if row:
                sql_text = str(row[0])
            cursor.close()
            conn.close()
        else:
            sql_text = args.sql_text

        if not sql_text:
            logger.error(f"SQL_ID {args.sql_id}의 SQL 텍스트를 찾을 수 없음")
            return

        trc_file = collector.collect_10053(args.sql_id, sql_text)
        if trc_file:
            logger.info(f"10053 트레이스 수집 완료: {trc_file}")
            # 자동 분석
            parsed = collector.parse_10053(trc_file)
            if parsed:
                report_dir = os.path.join(config["paths"]["report_output"], "10053")
                os.makedirs(report_dir, exist_ok=True)
                report_file = os.path.join(report_dir,
                    f"10053_{args.sql_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html")
                collector.generate_10053_report(parsed, report_file)
                logger.info(f"10053 리포트 생성: {report_file}")

    except Exception as e:
        logger.error(f"10053 트레이스 오류: {e}", exc_info=True)


def cmd_optimizer_analyze(args):
    """10053 트레이스 파일 분석"""
    from optimizer_trace import OptimizerTraceAnalyzer
    config = load_config(get_config_path())
    logger = setup_logger("optimizer_trace", config)

    try:
        analyzer = OptimizerTraceAnalyzer(config, logger)

        if args.file:
            trc_file = args.file
        else:
            # output/traces/ 에서 10053_ 파일 찾기
            trace_dir = config["paths"].get("trace_output", "output/traces")
            files = [f for f in os.listdir(trace_dir) if f.startswith("10053_") and f.endswith(".trc")]
            if not files:
                logger.info("분석할 10053 트레이스 파일 없음")
                return
            trc_file = os.path.join(trace_dir, sorted(files)[-1])
            logger.info(f"최신 10053 파일 분석: {trc_file}")

        parsed = analyzer.parse_10053(trc_file)
        if parsed:
            report_dir = os.path.join(config["paths"]["report_output"], "10053")
            os.makedirs(report_dir, exist_ok=True)
            base_name = os.path.splitext(os.path.basename(trc_file))[0]
            report_file = os.path.join(report_dir, f"{base_name}_report.html")
            analyzer.generate_10053_report(parsed, report_file)
            logger.info(f"10053 리포트 생성: {report_file}")

            # 요약 출력
            print(f"\n{'='*60}")
            print(f"  10053 Optimizer Trace Analysis")
            print(f"{'='*60}")
            if 'tables' in parsed:
                print(f"\n  테이블 통계: {len(parsed['tables'])}개")
                for t in parsed['tables']:
                    print(f"    - {t.get('table_name','?')}: rows={t.get('rows','?')}, blocks={t.get('blocks','?')}")
            if 'best_join_order' in parsed:
                print(f"\n  최적 조인 순서: {parsed['best_join_order']}")
            if 'total_cost' in parsed:
                print(f"  총 비용: {parsed['total_cost']}")
            if 'issues' in parsed and parsed['issues']:
                print(f"\n  ⚠️ 개선 포인트: {len(parsed['issues'])}건")
                for issue in parsed['issues']:
                    print(f"    - {issue}")
            print(f"\n  리포트: {report_file}")

    except Exception as e:
        logger.error(f"10053 분석 오류: {e}", exc_info=True)

def main():
    parser = argparse.ArgumentParser(
        description="Oracle SQL Tuning Automation",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
사용 예시:
  python main.py detect --dry-run     느린 SQL 감지 (테스트)
  python main.py run                  전체 파이프라인 실행
  python main.py report --daily       일간 리포트 생성
  python main.py export               Excel 내보내기
  python main.py status               상태 확인
  python main.py install-schedule     스케줄러 등록
        """,
    )
    subparsers = parser.add_subparsers(dest="command", help="실행할 명령")

    # detect
    p_detect = subparsers.add_parser("detect", help="Phase 1: 느린 SQL 감지")
    p_detect.add_argument("--dry-run", action="store_true", help="감지만 수행")

    # trace
    p_trace = subparsers.add_parser("trace", help="Phase 2: 트레이스 수집")
    p_trace.add_argument("--sql-id", type=str, help="단일 SQL ID")
    p_trace.add_argument("--file", type=str, help="감지 결과 JSON 파일")
    p_trace.add_argument("--wait", type=int, help="수집 대기 시간(초)")

    # analyze
    p_analyze = subparsers.add_parser("analyze", help="Phase 3: tkprof 분석")
    p_analyze.add_argument("--file", type=str, help="트레이스 파일 또는 디렉토리")

    # report
    p_report = subparsers.add_parser("report", help="Phase 4: 리포트 생성")
    p_report.add_argument("--daily", action="store_true", help="일간 리포트")
    p_report.add_argument("--weekly", action="store_true", help="주간 리포트")
    p_report.add_argument("--date", type=str, help="대상 날짜 (YYYY-MM-DD)")

    # run (전체 파이프라인)
    # target (특정 SQL_ID 분석)
    p_target = subparsers.add_parser("target", help="특정 SQL_ID 타겟 분석 (10046+tkprof+10053+Excel)")
    p_target.add_argument("--sql-id", type=str, required=True, help="분석할 SQL_ID")
    p_target.add_argument("--db-password", type=str, default=None, help="DB 비밀번호")
    p_target.add_argument("--skip-10046", action="store_true", help="10046 트레이스 스킵")
    p_target.add_argument("--skip-10053", action="store_true", help="10053 트레이스 스킵")
    p_target.add_argument("--skip-excel", action="store_true", help="Excel 리포트 스킵")

    p_run = subparsers.add_parser("run", help="전체 파이프라인 (Phase 1→2→3→4→10053→Excel)")
    p_run.add_argument("--skip-detect", action="store_true", help="Phase 1 스킵")
    p_run.add_argument("--skip-10053", action="store_true", help="10053 트레이스 스킵")
    p_run.add_argument("--skip-excel", action="store_true", help="Excel 리포트 스킵")
    p_run.add_argument("--db-password", type=str, default=None, help="DB 비밀번호")

    # cleanup
    p_cleanup = subparsers.add_parser("cleanup", help="오래된 파일 정리")

    # install-schedule
    p_schedule = subparsers.add_parser("install-schedule", help="스케줄러 등록")
    p_schedule.add_argument("--remove", action="store_true", help="스케줄러 등록 제거")

    # status
    p_status = subparsers.add_parser("status", help="상태 확인")

    # export
    p_export = subparsers.add_parser("export", help="Excel 내보내기")
    p_export.add_argument("--json", type=str, default=None,
                          help="AWR JSON 파일 또는 디렉토리 (기본: output/traces/)")
    p_export.add_argument("--detected", type=str, default=None,
                          help="Phase1 감지 결과 JSON 파일 또는 디렉토리")
    p_export.add_argument("--output", type=str, default=None,
                          help="출력 Excel 파일 경로 (기본: output/reports/sql_tuning_report_*.xlsx)")

    # 10053 optimizer trace
    p_opt_trace = subparsers.add_parser("optimizer-trace", help="10053 옵티마이저 트레이스 수집")
    p_opt_trace.add_argument("--sql-id", type=str, required=True, help="대상 SQL_ID")
    p_opt_trace.add_argument("--sql-text", type=str, default=None, help="SQL 텍스트 (생략 시 V$SQL에서 조회)")

    p_opt_analyze = subparsers.add_parser("optimizer-analyze", help="10053 트레이스 파일 분석")
    p_opt_analyze.add_argument("--file", type=str, default=None, help="10053 트레이스 파일 경로")

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        return

    commands = {
        "detect": cmd_detect,
        "trace": cmd_trace,
        "analyze": cmd_analyze,
        "report": cmd_report,
        "target": cmd_target,
        "run": cmd_run,
        "cleanup": cmd_cleanup,
        "install-schedule": cmd_install_schedule,
        "status": cmd_status,
        "export": cmd_export,
        "optimizer-trace": cmd_optimizer_trace,
        "optimizer-analyze": cmd_optimizer_analyze,
    }

    cmd_func = commands.get(args.command)
    if cmd_func:
        cmd_func(args)
    else:
        parser.print_help()


if __name__ == "__main__":

    main()
